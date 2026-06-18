"""
flowworkshop · dataframe_service.py
Speichert XLSX/CSV als echte SQL-Tabellen in PostgreSQL.
Ermoeglicht SQL-Abfragen und Statistik-Aggregationen durch das LLM.

Tabellen werden als workshop_df_{source_name} gespeichert.
"""
from __future__ import annotations
import csv
import io
import logging
import re
import unicodedata
from typing import Any

import pandas as pd
from rapidfuzz import fuzz
from sqlalchemy import text

from database import engine
from services.country_profiles import (
    country_code_for_bundesland,
    detect_country_code,
    get_country_name,
)

log = logging.getLogger(__name__)

_LEGAL_SUFFIXES = [
    r"gmbh\s*&\s*co\.?\s*kgaa",
    r"gmbh\s*&\s*co\.?\s*kg",
    r"gmbh\s*&\s*co\.?\s*ohg",
    r"ag\s*&\s*co\.?\s*kg",
    r"ag\s*&\s*co\.?\s*kgaa",
    r"ug\s*\(haftungsbeschr[aä]nkt\)",
    r"ggmbh",
    r"gmbh",
    r"kgaa",
    r"kg",
    r"ohg",
    r"ug",
    r"ag",
    r"se",
    r"gbr",
    r"mbh",
    r"e\.?\s*v\.?",
    r"e\.?\s*g\.?",
    r"inc\.?",
    r"ltd\.?",
    r"s\.?a\.?",
]

_LEGAL_RE = re.compile(r"\b(?:" + "|".join(_LEGAL_SUFFIXES) + r")\s*$", re.IGNORECASE)

_SEARCH_STOP_WORDS = frozenset({
    "und", "der", "die", "das", "fuer", "für", "von", "den", "dem", "des",
    "am", "im", "an", "zu", "zur", "zum", "bei", "mit", "the", "and", "for",
    "gmbh", "ggmbh", "mbh", "ltd", "inc", "kgaa",
})

_REFERENCE_PROFILE_PATTERNS: dict[str, dict[str, list[str]]] = {
    "sanctions": {
        "name": [
            r"^name$", r"entity_?name", r"subject_?name", r"full_?name",
            r"organisation_?name", r"organization_?name", r"company_?name",
            r"person_?name", r"designation",
        ],
        "alias": [r"aliases?", r"also_?known", r"\baka\b", r"other_?names?"],
        "projekt": [r"sanction_?program", r"sanction_?programme", r"program", r"programme", r"list_?name", r"regime"],
        "beschreibung": [r"sanction_?reasons?", r"reasons?", r"grounds?", r"remarks?", r"comment", r"description", r"title"],
        "aktenzeichen": [r"list_?id", r"entity_?id", r"subject_?id", r"reference", r"regulation", r"decision", r"listing_?id", r"^id$"],
        "standort": [r"address", r"city", r"town", r"place_?of_?birth", r"birth_?place", r"jurisdiction"],
        "country": [r"country", r"citizenship", r"nationality", r"jurisdiction_?country"],
        "status": [r"listing_?status", r"sanction_?status", r"status", r"active", r"state"],
        "date": [r"listing_?date", r"date_?listed", r"effective_?date", r"start_?date", r"date"],
    },
    "tam": {
        "name": [
            r"beneficiary_?name", r"name_?of_?the_?beneficiary", r"undertaking",
            r"recipient_?name", r"company_?name", r"^beneficiary$",
        ],
        "projekt": [r"aid_?measure_?title", r"measure_?title", r"scheme_?title", r"aid_?scheme", r"title", r"measure", r"scheme"],
        "beschreibung": [r"objectives?", r"aid_?instrument", r"instrument", r"sector(_?nace)?", r"economic_?activity", r"beneficiary_?type", r"entrusted_?entity", r"remarks?"],
        "aktenzeichen": [r"sa_?number", r"ref_?no", r"reference_?number", r"national_?id", r"case_?number", r"measure_?id"],
        "standort": [r"region", r"city", r"address", r"postal", r"postcode", r"zip", r"location"],
        "country": [r"member_?state", r"country"],
        "status": [r"is_?late", r"late_?publication", r"delay", r"publication_?status", r"status"],
        "authority": [r"granting_?authority", r"awarding_?authority", r"authority", r"provider"],
        "amount": [r"aid_?element", r"nominal_?amount", r"gross_?grant", r"grant_?equivalent", r"amount", r"value"],
        "programme": [r"program(me)?", r"regime", r"objective"],
        "date": [r"date_?granted", r"grant_?date", r"published_?date", r"publication_?date", r"award_?date", r"date"],
    },
    "state_aid": {
        "name": [r"beneficiary_?name", r"recipient_?name", r"undertaking", r"company_?name", r"^beneficiary$", r"^name$"],
        "projekt": [r"aid_?measure_?title", r"measure_?title", r"scheme_?title", r"measure_?name", r"title", r"scheme"],
        "beschreibung": [r"objectives?", r"objective", r"aid_?instrument", r"instrument", r"legal_?basis", r"sector", r"activity", r"description", r"summary"],
        "aktenzeichen": [r"sa_?number", r"case_?number", r"measure_?id", r"reference_?number", r"notification_?number", r"^id$"],
        "standort": [r"region", r"city", r"address", r"postal", r"postcode", r"zip", r"location"],
        "country": [r"member_?state", r"country"],
        "status": [r"status", r"phase", r"decision"],
        "authority": [r"granting_?authority", r"awarding_?authority", r"authority", r"managing_?authority", r"provider"],
        "amount": [r"aid_?element", r"nominal_?amount", r"gross_?grant", r"amount", r"value", r"budget"],
        "programme": [r"program(me)?", r"fund", r"objective", r"priority"],
        "date": [r"date_?granted", r"grant_?date", r"published_?date", r"decision_?date", r"approval_?date", r"date"],
    },
    "cohesio": {
        "name": [r"beneficiary", r"final_?recipient", r"recipient_?name", r"company_?name", r"organisation", r"organization", r"undertaking", r"^name$"],
        "projekt": [r"operation_?name", r"operation_?title", r"project_?title", r"project_?name", r"title", r"operation"],
        "beschreibung": [r"program(me)?", r"fund", r"policy_?objective", r"intervention_?field", r"description", r"summary", r"theme", r"category"],
        "aktenzeichen": [r"operation_?id", r"project_?id", r"reference", r"cci", r"code", r"^id$"],
        "standort": [r"region", r"city", r"municipality", r"nuts", r"location"],
        "country": [r"country", r"member_?state"],
        "status": [r"implementation_?status", r"status", r"phase"],
        "amount": [r"eu_?contribution", r"union_?contribution", r"total_?budget", r"budget", r"total_?cost", r"amount"],
        "programme": [r"program(me)?", r"fund", r"priority", r"objective"],
        "date": [r"start_?date", r"end_?date", r"approval_?date", r"completion_?date", r"date"],
    },
}

_REFERENCE_CONTEXT_LABELS = {
    "alias": "Alias",
    "authority": "Stelle",
    "amount": "Volumen",
    "programme": "Programm",
    "beschreibung": "Kontext",
    "date": "Datum",
}

_REFERENCE_CONTEXT_ORDER = {
    "sanctions": ["alias", "beschreibung", "date"],
    "tam": ["amount", "authority", "programme", "beschreibung", "date"],
    "state_aid": ["amount", "authority", "programme", "beschreibung", "date"],
    "cohesio": ["programme", "amount", "beschreibung", "date"],
    "other": ["beschreibung", "programme", "amount", "authority", "date"],
}


def _safe_table_name(source: str) -> str:
    """Erzeugt einen sicheren SQL-Tabellennamen aus dem Source-Label."""
    name = re.sub(r"[^a-zA-Z0-9_]", "_", source.lower()).strip("_")
    return f"workshop_df_{name}"


def _quote_ident(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def _parse_numeric(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text_value = str(value).strip()
    if not text_value:
        return None
    cleaned = text_value.replace("\xa0", " ").replace("EUR", "").replace("€", "")
    cleaned = cleaned.replace(" ", "")
    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def _format_eur(value: float | None) -> str:
    if value is None:
        return "k.A."
    return f"{value:,.0f}".replace(",", ".") + " €"


def _format_eur_compact(value: float | None) -> str:
    """Kompakte, gut lesbare Euro-Darstellung (Mrd./Mio./Tsd.) für Ranglisten.

    Beispiele: 2_318_000_000 -> '2,32 Mrd. €', 602_000_000 -> '602 Mio. €'.
    0/None ergibt einen ehrlichen Hinweis statt '0 €', weil viele
    Transparenzlisten keine Beträge ausweisen.
    """
    if value is None or value == 0:
        return "keine Betragsangabe"
    abs_v = abs(value)
    if abs_v >= 1_000_000_000:
        return f"{value / 1_000_000_000:.2f}".replace(".", ",") + " Mrd. €"
    if abs_v >= 1_000_000:
        return f"{value / 1_000_000:.0f}" + " Mio. €"
    if abs_v >= 10_000:
        return f"{value / 1_000:.0f}" + " Tsd. €"
    return _format_eur(value)


def _de_int(value: int | float | None) -> str:
    """Formatiert eine Ganzzahl mit deutschem Tausenderpunkt (z.B. 4501 -> '4.501').

    None wird als '0' ausgegeben, damit f-Strings sauber bleiben.
    """
    if value is None:
        return "0"
    try:
        return f"{int(value):,}".replace(",", ".")
    except (TypeError, ValueError):
        return str(value)


def _normalize_search_text(value: Any) -> str:
    text_value = str(value or "").strip().lower()
    text_value = re.sub(r"\s+", " ", text_value)
    return text_value


def _compact_search_text(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", _normalize_search_text(value))


def _strip_legal_suffix(value: Any) -> str:
    normalized = _normalize_search_text(value)
    if not normalized:
        return ""
    cleaned = _LEGAL_RE.sub("", normalized).strip()
    cleaned = re.sub(r"\s*&\s*(co\.?)?\s*$", "", cleaned, flags=re.IGNORECASE).strip()
    return cleaned.rstrip(",;. ")


def _split_word_separators(value: str) -> str:
    """Ersetzt Bindestrich und Slash durch Leerzeichen, damit
    'Fraunhofer-Gesellschaft' und 'Fraunhofer Gesellschaft' identisch
    behandelt werden. Wird vor Tokenize/Substring-Matches angewendet.
    """
    if not value:
        return value
    return value.replace("-", " ").replace("/", " ")


def _tokenize_search_text(value: Any) -> list[str]:
    normalized = _strip_legal_suffix(value) or _normalize_search_text(value)
    if not normalized:
        return []
    # Hyphen + Slash als Wort-Trenner behandeln, damit
    # "Fraunhofer-Gesellschaft" zu Tokens [fraunhofer, gesellschaft] wird.
    normalized = _split_word_separators(normalized)
    tokens = re.split(r"[\s,;|()]+", normalized)
    return [token for token in tokens if len(token) >= 3 and token not in _SEARCH_STOP_WORDS]


def _search_word_present(token: str, text_value: str) -> bool:
    # Hyphen/Slash → Space, damit \b-Match auch ueber Hyphen-Grenzen klappt
    haystack = _split_word_separators(text_value)
    if re.search(r"\b" + re.escape(token) + r"\b", haystack, re.IGNORECASE):
        return True
    if len(token) >= 6 and re.search(r"\b" + re.escape(token[:6]), haystack, re.IGNORECASE):
        return True
    return False


def _score_search_value(value: Any, query: str) -> int:
    normalized = _normalize_search_text(value)
    if not normalized or not query:
        return 0

    # Vereinheitlichte Variante (Hyphen/Slash → Space) fuer Substring-
    # und Word-Matches. Bewahrt 'normalized' fuer reine Equality-Tests.
    normalized_split = _split_word_separators(normalized)
    query_split = _split_word_separators(query)

    stripped_normalized = _strip_legal_suffix(normalized)
    stripped_query = _strip_legal_suffix(query)
    stripped_normalized_split = _split_word_separators(stripped_normalized)
    stripped_query_split = _split_word_separators(stripped_query)
    query_tokens = _tokenize_search_text(query)

    if normalized == query:
        return 140
    if stripped_query and stripped_normalized == stripped_query:
        return 132
    if normalized.startswith(query) or normalized_split.startswith(query_split):
        return 115
    if stripped_query and (
        stripped_normalized.startswith(stripped_query)
        or stripped_normalized_split.startswith(stripped_query_split)
    ):
        return 109
    if query_tokens and all(_search_word_present(token, normalized) for token in query_tokens):
        return 104
    if all(part in normalized_split for part in query_split.split()):
        return 92
    if stripped_query and (
        stripped_query in stripped_normalized
        or stripped_query_split in stripped_normalized_split
    ):
        return 88
    if query in normalized or query_split in normalized_split:
        return 78

    compact_query = _compact_search_text(query)
    compact_value = _compact_search_text(normalized)
    if compact_query and compact_query in compact_value:
        return 68
    return 0


# ── rapidfuzz-Scoring fuer Beguenstigtenliste ────────────────────────────────
# Ziel: gleiche Skala (0..100) wie State-Aid-Suche
# (services/state_aid_service.fuzzy_match_company), damit Workshop-Module
# konsistent „fuehlen" und Confidence-Klassen vergleichbar sind.

_ACCENT_TABLE = str.maketrans({
    "ä": "ae", "ö": "oe", "ü": "ue", "ß": "ss",
    "Ä": "ae", "Ö": "oe", "Ü": "ue",
    "á": "a", "à": "a", "â": "a", "ã": "a", "å": "a",
    "é": "e", "è": "e", "ê": "e", "ë": "e",
    "í": "i", "ì": "i", "î": "i", "ï": "i",
    "ó": "o", "ò": "o", "ô": "o", "õ": "o",
    "ú": "u", "ù": "u", "û": "u",
    "ç": "c", "ñ": "n", "ý": "y",
    "ł": "l", "ń": "n", "ś": "s", "ź": "z", "ż": "z",
    "č": "c", "š": "s", "ž": "z", "đ": "d",
})


def _normalize_for_rapidfuzz(value: Any) -> str:
    """Vergleichsform fuer rapidfuzz: lowercase, ohne Akzente/Umlaute,
    Hyphen/Slash → Space, Whitespace kompakt.

    Bewusst KEIN Strippen von Rechtsform-Suffixen (RapidFuzz token_set_ratio
    blendet Tokens, die nur im einen oder anderen String vorkommen, ohnehin
    aus). So bleibt die Funktion auch fuer Identifier-Felder (Aktenzeichen)
    nutzbar, in denen Suffixe nicht vorkommen.
    """
    if not value:
        return ""
    text_value = str(value).translate(_ACCENT_TABLE).casefold()
    text_value = text_value.replace("-", " ").replace("/", " ")
    text_value = re.sub(r"\s+", " ", text_value).strip()
    return text_value


def _rapidfuzz_score(value: Any, query: str) -> float:
    """Liefert max(token_set_ratio, WRatio) auf Skala 0..100.

    - Akzent-/Hyphen-Normalisierung vorab (siehe _normalize_for_rapidfuzz),
      damit 'Fraunhofer-Gesellschaft' und 'Fraunhofer Gesellschaft' identisch
      bewertet werden.
    - Exact-Match-Boost: Wenn die normalisierten Strings identisch sind,
      wird sofort 100.0 zurueckgegeben (analog dem Score-140-Sonderfall im
      Legacy-Scorer, hier auf 0..100-Skala normiert).
    - Leere Eingaben → 0.0.
    """
    if not value or not query:
        return 0.0
    norm_value = _normalize_for_rapidfuzz(value)
    norm_query = _normalize_for_rapidfuzz(query)
    if not norm_value or not norm_query:
        return 0.0
    if norm_value == norm_query:
        return 100.0
    tsr = fuzz.token_set_ratio(norm_query, norm_value)
    wr = fuzz.WRatio(norm_query, norm_value)
    return float(max(tsr, wr))


def _match_confidence(score: float) -> str:
    """Konfidenz-Klasse analog services/state_aid_service._confidence:
    >=97 exact, >=90 high, >=80 medium, sonst low.
    """
    if score >= 97:
        return "exact"
    if score >= 90:
        return "high"
    if score >= 80:
        return "medium"
    return "low"


def _adaptive_min_score(query: str) -> float:
    """Heuristik analog routers/state_aid._adaptive_min_score:
    1 Token → 80, 2 Tokens → 70, ≥3 Tokens → 60.
    """
    n_tokens = len((query or "").split())
    if n_tokens >= 3:
        return 60.0
    if n_tokens == 2:
        return 70.0
    return 80.0


def ingest_dataframe(
    file_bytes: bytes,
    filename: str,
    source: str,
    sheet_name: str | int | None = 0,
    dataset_group: str = "generic",
    registry_type: str | None = None,
) -> dict:
    """
    Liest eine XLSX/XLS/CSV-Datei als DataFrame ein und speichert sie
    als SQL-Tabelle. Bestehende Tabelle wird ersetzt.

    Returns:
        {"table_name": str, "rows": int, "columns": list[str], "source": str}
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext in ("xlsx", "xls", "xlsm"):
        # Immer smart header detection (EFRE-Tabellen haben oft Titelzeilen)
        df = _read_excel_smart(file_bytes, ext, sheet_name)
    elif ext == "csv":
        df = _read_csv_smart(file_bytes)
    else:
        raise ValueError(f"DataFrame-Ingest nur fuer XLSX/XLS/CSV, nicht '{ext}'")

    # Spalten bereinigen und Kollisionen aufloesen
    df.columns = _make_unique_column_names([str(c) for c in df.columns])

    # Leere Zeilen/Spalten entfernen
    df = df.dropna(how="all").dropna(axis=1, how="all")

    # Unnamed-Spalten entfernen
    df = df[[c for c in df.columns if not c.startswith("unnamed")]]

    if df.empty:
        return {"table_name": "", "rows": 0, "columns": [], "source": source}

    table_name = _safe_table_name(source)

    # Metadaten aus Titelzeilen extrahieren (Bundesland, Fonds, Periode)
    metadata = {}
    if ext in ("xlsx", "xls", "xlsm", "csv"):
        metadata = _detect_metadata(file_bytes, ext, sheet_name, source=filename)

    # In PostgreSQL speichern (replace = DROP + CREATE)
    df.to_sql(table_name, engine, if_exists="replace", index=False)

    # Metadaten in separater Tabelle speichern
    _save_metadata(
        table_name,
        source,
        metadata,
        dataset_group=dataset_group,
        registry_type=registry_type,
        filename=filename,
    )

    log.info("DataFrame %s: %d Zeilen, %d Spalten → %s (%s)",
             source, len(df), len(df.columns), table_name, metadata)

    return {
        "table_name": table_name,
        "rows": len(df),
        "columns": list(df.columns),
        "source": source,
        "dtypes": {c: str(df[c].dtype) for c in df.columns},
        "metadata": metadata,
        "dataset_group": dataset_group,
        "registry_type": registry_type,
    }


def _read_excel_smart(file_bytes: bytes, ext: str, sheet_name: str | int | None = 0) -> pd.DataFrame:
    """
    Versucht die Header-Zeile automatisch zu finden.
    Typisch bei EFRE-Transparenzlisten: 3-6 leere/Titel-Zeilen vor dem Header.
    Heuristik: Erste Zeile mit >= 3 nicht-leeren Zellen UND mindestens eine
    Zelle die wie ein Spaltenname aussieht (kurz, alphabetisch).
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    # Richtiges Blatt auswaehlen
    if isinstance(sheet_name, str) and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    elif isinstance(sheet_name, int) and sheet_name < len(wb.worksheets):
        ws = wb.worksheets[sheet_name]
    else:
        ws = wb.worksheets[0]

    best_row = 0
    candidates = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 30:
            break
        cells = [c for c in row if c is not None]
        non_empty = [str(c).strip() for c in cells if str(c).strip()]

        if len(non_empty) < 2:
            continue

        # Header-Heuristik: Zellen mit Buchstaben zaehlen
        text_cells = sum(1 for c in non_empty if any(ch.isalpha() for ch in str(c)))
        # Reine Zahlen-Zellen (typisch fuer Datenzeilen)
        pure_numbers = sum(1 for c in non_empty if _is_number(c))

        # Header hat typischerweise mehr Text als Zahlen
        # Datenzeilen haben viele Zahlen
        text_ratio = text_cells / max(len(non_empty), 1)

        candidates.append((i, len(non_empty), text_ratio, text_cells, pure_numbers))

    if candidates:
        # Sortiere: Bevorzuge Zeilen mit hoher Text-Ratio und vielen Zellen
        # Erster Kandidat mit text_ratio > 0.5 und >= 3 Zellen
        for row_idx, count, ratio, texts, nums in candidates:
            if count >= 3 and ratio >= 0.4:
                best_row = row_idx
                break
        else:
            # Fallback: Zeile mit den meisten Zellen
            best_row = max(candidates, key=lambda x: (x[2], x[1]))[0]

    wb.close()

    df = pd.read_excel(
        io.BytesIO(file_bytes),
        header=best_row,
        sheet_name=sheet_name,
        engine="openpyxl",
    )
    return df


def _read_csv_smart(file_bytes: bytes) -> pd.DataFrame:
    """
    Erkennt Delimiter und Header-Zeile fuer Transparenzlisten im CSV-Format.
    Behandelt Titelzeilen vor dem Header und zweisprachige Doppel-Header.
    """
    text = ""
    for enc in ("utf-8", "cp1252", "latin-1"):
        try:
            text = file_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if not text:
        text = file_bytes.decode("utf-8", errors="replace")

    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
        sep = dialect.delimiter
    except csv.Error:
        sep = ";"

    raw = pd.read_csv(
        io.StringIO(text),
        sep=sep,
        header=None,
        dtype=str,
        engine="python",
        keep_default_na=False,
    )

    best_row = 0
    candidates: list[tuple[int, int, float]] = []
    for idx in range(min(len(raw), 25)):
        values = [str(v).strip() for v in raw.iloc[idx].tolist()]
        non_empty = [v for v in values if v]
        if len(non_empty) < 2:
            continue
        text_cells = sum(1 for v in non_empty if any(ch.isalpha() for ch in v))
        text_ratio = text_cells / max(len(non_empty), 1)
        candidates.append((idx, len(non_empty), text_ratio))

    if candidates:
        for idx, count, ratio in candidates:
            if count >= 3 and ratio >= 0.6:
                best_row = idx
                break
        else:
            best_row = max(candidates, key=lambda item: (item[2], item[1]))[0]

    # Manche Transparenzlisten haben DE-Header + maschinenlesbaren EN-Header direkt darunter.
    if best_row + 1 < len(raw):
        next_values = [str(v).strip() for v in raw.iloc[best_row + 1].tolist()]
        next_non_empty = [v for v in next_values if v]
        snake_like = sum(1 for v in next_non_empty if re.fullmatch(r"[a-z0-9_]+", v))
        if next_non_empty and snake_like >= max(3, len(next_non_empty) // 2):
            best_row += 1

    return pd.read_csv(
        io.StringIO(text),
        sep=sep,
        header=best_row,
        engine="python",
    )


def _clean_column_name(name: str) -> str:
    """Bereinigt Spaltennamen fuer SQL und behaelt genug Kontext fuer mehrzeilige Header."""
    name = name.strip()
    if "\n" in name:
        name = " ".join(part.strip() for part in name.splitlines() if part.strip())
    name = name.lower()
    name = re.sub(r"\s+", "_", name)
    name = re.sub(r"[^a-zA-Z0-9_äöüß]", "", name)
    name = name.strip("_")
    if not name or name[0].isdigit():
        name = "col_" + name
    return name[:63]  # PostgreSQL Limit


def _make_unique_column_names(names: list[str]) -> list[str]:
    """Bereinigt Spaltennamen und haengt bei Kollisionen stabile Suffixe an."""
    counts: dict[str, int] = {}
    unique_names: list[str] = []

    for raw_name in names:
        base = _clean_column_name(raw_name)
        next_index = counts.get(base, 0)
        counts[base] = next_index + 1

        if next_index == 0:
            unique_names.append(base)
            continue

        suffix = f"_{next_index + 1}"
        trimmed = base[: max(1, 63 - len(suffix))]
        unique_names.append(f"{trimmed}{suffix}")

    return unique_names


def _is_number(s: str) -> bool:
    """Prüft ob ein String eine Zahl ist."""
    try:
        float(str(s).replace(",", ".").replace(" ", ""))
        return True
    except (ValueError, TypeError):
        return False


# ── Metadaten-Erkennung ──────────────────────────────────────────────────────

BUNDESLAENDER = [
    # Deutschland
    "Baden-Württemberg", "Bayern", "Berlin", "Brandenburg", "Bremen",
    "Hamburg", "Hessen", "Mecklenburg-Vorpommern", "Niedersachsen",
    "Nordrhein-Westfalen", "Rheinland-Pfalz", "Saarland", "Sachsen",
    "Sachsen-Anhalt", "Schleswig-Holstein", "Thüringen",
    # Österreich
    "Burgenland", "Kärnten", "Niederösterreich", "Oberösterreich",
    "Salzburg", "Steiermark", "Tirol", "Vorarlberg", "Wien",
]

BUNDESLAND_ALIASES = {
    # Deutschland
    "Baden-Württemberg": ["Baden-Württemberg", "Baden-Wuerttemberg"],
    "Bayern": ["Bayern", "Freistaat Bayern"],
    "Berlin": ["Berlin"],
    "Brandenburg": ["Brandenburg"],
    "Bremen": ["Bremen"],
    "Hamburg": ["Hamburg"],
    "Hessen": ["Hessen"],
    "Mecklenburg-Vorpommern": ["Mecklenburg-Vorpommern", "Mecklenburg Vorpommern"],
    "Niedersachsen": ["Niedersachsen"],
    "Nordrhein-Westfalen": ["Nordrhein-Westfalen", "Nordrhein Westfalen", "NRW"],
    "Rheinland-Pfalz": ["Rheinland-Pfalz", "Rheinland Pfalz"],
    "Saarland": ["Saarland"],
    "Sachsen": ["Sachsen", "Freistaat Sachsen"],
    "Sachsen-Anhalt": ["Sachsen-Anhalt", "Sachsen Anhalt"],
    "Schleswig-Holstein": ["Schleswig-Holstein", "Schleswig Holstein"],
    "Thüringen": ["Thüringen", "Thueringen", "Freistaat Thüringen"],
    # Österreich
    "Burgenland": ["Burgenland"],
    "Kärnten": ["Kärnten", "Kaernten", "Carinthia"],
    "Niederösterreich": ["Niederösterreich", "Niederoesterreich", "Lower Austria"],
    "Oberösterreich": ["Oberösterreich", "Oberoesterreich", "Upper Austria"],
    "Salzburg": ["Salzburg", "Land Salzburg"],
    "Steiermark": ["Steiermark", "Styria"],
    "Tirol": ["Tirol", "Tyrol"],
    "Vorarlberg": ["Vorarlberg"],
    "Wien": ["Wien", "Vienna"],
}

FONDS = [
    "EFRE/JTF", "ESF+/JTF",
    "EFRE", "ESF+", "ESF", "ELER", "EMFAF", "ERDF", "JTF", "AMIF", "ISF", "REACT-EU",
]

PERIODEN = ["2014-2020", "2021-2027", "2014–2020", "2021–2027",
            "21-27", "14-20", "2028-2034"]


def _detect_metadata(
    file_bytes: bytes,
    ext: str,
    sheet_name: str | int | None = 0,
    source: str | None = None,
) -> dict:
    """
    Erkennt Bundesland, Fonds und Foerderperiode aus den Titelzeilen einer XLSX.
    Liest die ersten 15 Zeilen VOR dem Daten-Header.
    Falls nichts erkannt wird, wird der Dateiname (source) als Fallback herangezogen.
    """
    title_text = ""
    if ext in ("xlsx", "xls", "xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

        if isinstance(sheet_name, str) and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        elif isinstance(sheet_name, int) and sheet_name < len(wb.worksheets):
            ws = wb.worksheets[sheet_name]
        else:
            ws = wb.worksheets[0]

        # Alle Texte aus den ersten 15 Zeilen sammeln
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 15:
                break
            for c in row:
                if c is not None:
                    title_text += " " + str(c)
        wb.close()
    else:
        for enc in ("utf-8", "cp1252", "latin-1"):
            try:
                title_text = file_bytes[:16000].decode(enc)
                break
            except UnicodeDecodeError:
                continue

    result = {
        "bundesland": None,
        "fonds": None,
        "periode": None,
        "country_code": None,
        "country_name": None,
    }

    title_text_lower = title_text.lower()
    source_lower = (source or "").lower()
    normalized_title = _normalize_lookup_text(title_text)
    normalized_source = _normalize_lookup_text(source or "")

    # Bundesland erkennen: Dateiname hat Vorrang. Titel-Fallback nur, wenn der
    # Dateiname kein Land verraet — sonst wuerden Datenzeilen einer Gesamtliste
    # ein zufaellig zuerst auftauchendes Bundesland setzen (z.B. AT-Gesamtliste
    # mit erstem Datensatz "Niederoesterreich").
    source_state = _detect_bundesland_from_text(normalized_source)
    source_country_hint = detect_country_code(source)
    if source_state:
        result["bundesland"] = source_state
    elif not source_country_hint:
        title_state = _detect_bundesland_from_text(normalized_title)
        if title_state:
            result["bundesland"] = title_state

    # Country-Code aus Bundesland (gewinnt eindeutig), sonst aus Datei-/Titel-Text
    country_code = country_code_for_bundesland(result["bundesland"])
    if not country_code:
        country_code = detect_country_code(source, title_text)
    if country_code:
        result["country_code"] = country_code
        result["country_name"] = get_country_name(country_code)

    # Fonds erkennen
    for f in FONDS:
        if f.lower() in title_text_lower or f in title_text:
            result["fonds"] = f.upper()
            break

    # Periode erkennen
    for p in PERIODEN:
        if p in title_text:
            # Normalisieren
            result["periode"] = p.replace("–", "-")
            if len(result["periode"]) <= 5:
                result["periode"] = "20" + result["periode"]
            break

    # Fallback: aus Spaltenwerten (Fonds-Spalte)
    if not result["fonds"]:
        # Suche nach EFRE/ESF in den Daten
        for f in FONDS:
            if f.lower() in title_text_lower:
                result["fonds"] = f
                break

    # Fallback: Bundesland und Fonds aus Dateinamen erkennen
    # Greift z.B. bei "transparenzliste_sachsen_efre.xlsx" oder "Begünstigte_NRW_2021-2027.xlsx"
    if source:
        source_state = _detect_bundesland_from_text(normalized_source)
        if source_state:
            result["bundesland"] = source_state

        for f in FONDS:
            if f.lower() in source_lower:
                result["fonds"] = f.upper()
                break

        for p in PERIODEN:
            p_normalized = p.replace("–", "-")
            if p_normalized in source or p_normalized.replace("-", "_") in source_lower:
                result["periode"] = p_normalized
                if len(result["periode"]) <= 5:
                    result["periode"] = "20" + result["periode"]
                break

    return result


def _normalize_lookup_text(value: str) -> str:
    """Normalisiert Text fuer robuste Laender- und Dateinamen-Erkennung."""
    text = str(value or "").lower()
    text = (
        text.replace("ä", "ae")
        .replace("ö", "oe")
        .replace("ü", "ue")
        .replace("ß", "ss")
    )
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[_/\\.-]+", " ", text)
    text = re.sub(r"[^a-z0-9 ]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _detect_bundesland_from_text(normalized_text: str) -> str | None:
    """Findet das passendste Bundesland in normalisiertem Text."""
    if not normalized_text:
        return None

    candidates: list[tuple[int, str]] = []
    for canonical, aliases in BUNDESLAND_ALIASES.items():
        for alias in aliases:
            normalized_alias = _normalize_lookup_text(alias)
            if normalized_alias and normalized_alias in normalized_text:
                candidates.append((len(normalized_alias), canonical))

    if not candidates:
        return None

    candidates.sort(key=lambda item: item[0], reverse=True)
    return candidates[0][1]


def _ensure_metadata_table():
    """Erstellt die Metadaten-Tabelle falls nicht vorhanden."""
    with engine.connect() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS workshop_df_metadata (
                table_name TEXT PRIMARY KEY,
                source TEXT NOT NULL,
                bundesland TEXT,
                fonds TEXT,
                periode TEXT,
                row_count INT DEFAULT 0,
                is_beneficiary BOOLEAN DEFAULT FALSE,
                dataset_group TEXT DEFAULT 'generic',
                registry_type TEXT,
                filename TEXT,
                country_code TEXT,
                country_name TEXT
            )
        """))
        conn.execute(text("""
            ALTER TABLE workshop_df_metadata
            ADD COLUMN IF NOT EXISTS dataset_group TEXT DEFAULT 'generic'
        """))
        conn.execute(text("""
            ALTER TABLE workshop_df_metadata
            ADD COLUMN IF NOT EXISTS registry_type TEXT
        """))
        conn.execute(text("""
            ALTER TABLE workshop_df_metadata
            ADD COLUMN IF NOT EXISTS filename TEXT
        """))
        conn.execute(text("""
            ALTER TABLE workshop_df_metadata
            ADD COLUMN IF NOT EXISTS country_code TEXT
        """))
        conn.execute(text("""
            ALTER TABLE workshop_df_metadata
            ADD COLUMN IF NOT EXISTS country_name TEXT
        """))
        # Bestehende deutsche Quellen ohne country_code automatisch auf DE setzen
        conn.execute(text("""
            UPDATE workshop_df_metadata
            SET country_code = 'DE',
                country_name = 'Deutschland'
            WHERE country_code IS NULL
              AND bundesland IN (
                'Baden-Württemberg', 'Bayern', 'Berlin', 'Brandenburg', 'Bremen',
                'Hamburg', 'Hessen', 'Mecklenburg-Vorpommern', 'Niedersachsen',
                'Nordrhein-Westfalen', 'Rheinland-Pfalz', 'Saarland', 'Sachsen',
                'Sachsen-Anhalt', 'Schleswig-Holstein', 'Thüringen'
              )
        """))
        conn.execute(text("""
            UPDATE workshop_df_metadata
            SET country_code = 'AT',
                country_name = 'Österreich'
            WHERE country_code IS NULL
              AND bundesland IN (
                'Burgenland', 'Kärnten', 'Niederösterreich', 'Oberösterreich',
                'Salzburg', 'Steiermark', 'Tirol', 'Vorarlberg', 'Wien'
              )
        """))
        conn.commit()


def _save_metadata(
    table_name: str,
    source: str,
    metadata: dict,
    dataset_group: str = "generic",
    registry_type: str | None = None,
    filename: str | None = None,
):
    """Speichert Metadaten zu einer DataFrame-Tabelle."""
    _ensure_metadata_table()

    # Ist es ein Beguenstigtenverzeichnis? (Hat Standort/Ort/PLZ-Spalte)
    from services.geocoding_service import detect_columns
    try:
        cols = detect_columns(source)
        is_beneficiary = bool(cols.get("standort") or cols.get("ort") or cols.get("plz"))
    except Exception:
        is_beneficiary = False
    if dataset_group == "beneficiary":
        is_beneficiary = True

    with engine.connect() as conn:
        # Zeilenanzahl
        try:
            count = conn.execute(text(f'SELECT COUNT(*) FROM "{table_name}"')).scalar()
        except Exception:
            count = 0

        # Country-Code aus Metadaten oder erschlossen aus Bundesland
        country_code = metadata.get("country_code")
        if not country_code:
            country_code = country_code_for_bundesland(metadata.get("bundesland"))
        country_name = metadata.get("country_name") or get_country_name(country_code)

        conn.execute(text("""
            INSERT INTO workshop_df_metadata (
                table_name, source, bundesland, fonds, periode, row_count,
                is_beneficiary, dataset_group, registry_type, filename,
                country_code, country_name
            )
            VALUES (:t, :s, :bl, :f, :p, :c, :b, :g, :rt, :fn, :cc, :cn)
            ON CONFLICT (table_name) DO UPDATE SET
                source = :s, bundesland = :bl, fonds = :f, periode = :p,
                row_count = :c, is_beneficiary = :b,
                dataset_group = :g, registry_type = :rt, filename = :fn,
                country_code = :cc, country_name = :cn
        """), {
            "t": table_name, "s": source,
            "bl": metadata.get("bundesland"),
            "f": metadata.get("fonds"),
            "p": metadata.get("periode"),
            "c": count,
            "b": is_beneficiary,
            "g": dataset_group,
            "rt": registry_type,
            "fn": filename,
            "cc": country_code,
            "cn": country_name,
        })
        conn.commit()


def get_beneficiary_sources(country_code: str | None = None) -> list[dict]:
    """Gibt alle DataFrame-Tabellen zurueck die Beguenstigtenverzeichnisse sind.

    Wenn country_code gesetzt ist, werden nur die Quellen dieses Landes
    zurueckgegeben (DE/AT). Quellen ohne hinterlegten country_code, deren
    Bundesland aber zu einem bekannten Land gehoert, werden mitgeliefert.
    """
    _ensure_metadata_table()
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT table_name, source, bundesland, fonds, periode, row_count,
                   dataset_group, registry_type, filename, country_code, country_name
            FROM workshop_df_metadata
            WHERE dataset_group = 'beneficiary'
               OR (is_beneficiary = TRUE AND COALESCE(dataset_group, 'generic') IN ('generic', 'beneficiary'))
            ORDER BY country_code NULLS LAST, bundesland, source
        """)).fetchall()

    sources = []
    target_country = country_code.upper() if country_code else None
    for r in rows:
        stored_cc = (r[9] or "").upper() or None
        # Fallback: aus Bundesland erschliessen
        derived_cc = stored_cc or country_code_for_bundesland(r[2])
        derived_name = r[10] or get_country_name(derived_cc)
        if target_country and derived_cc != target_country:
            continue
        sources.append({
            "table_name": r[0],
            "source": r[1],
            "bundesland": r[2],
            "fonds": r[3],
            "periode": r[4],
            "row_count": r[5],
            "dataset_group": r[6],
            "registry_type": r[7],
            "filename": r[8],
            "country_code": derived_cc,
            "country_name": derived_name,
        })
    return sources


def list_reference_registry_sources() -> list[dict]:
    """Gibt importierte Referenzregister fuer Unternehmensabgleiche zurueck."""
    _ensure_metadata_table()
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT table_name, source, row_count, dataset_group, registry_type, filename
            FROM workshop_df_metadata
            WHERE dataset_group = 'reference_registry'
               OR registry_type IS NOT NULL
            ORDER BY registry_type, source
        """)).fetchall()
    return [
        {
            "table_name": r[0],
            "source": r[1],
            "row_count": r[2],
            "dataset_group": r[3],
            "registry_type": r[4],
            "filename": r[5],
        }
        for r in rows
    ]


def query_dataframe(source: str, sql_query: str) -> list[dict]:
    """
    Fuehrt eine SQL-Abfrage auf einer DataFrame-Tabelle aus.
    SICHERHEIT: Nur SELECT auf die spezifische Tabelle, kein DML/DDL.
    """
    table_name = _safe_table_name(source)
    cleaned = sql_query.strip()
    upper = cleaned.upper()

    if not upper.startswith("SELECT"):
        raise ValueError("Nur SELECT-Abfragen erlaubt.")
    if ";" in cleaned:
        raise ValueError("Mehrere Statements (;) sind nicht erlaubt.")

    blocked = [
        "INSERT",
        "UPDATE",
        "DELETE",
        "DROP",
        "ALTER",
        "CREATE",
        "TRUNCATE",
        "GRANT",
        "REVOKE",
        "EXEC",
        "EXECUTE",
        "INTO",
        "COPY",
        "UNION",
        "INTERSECT",
        "EXCEPT",
        "JOIN",
        "WITH",
        # DoS-Vektoren in Postgres: Sleep blockiert den Worker, lo_*/pg_read_file
        # zielen auf Filesystem, pg_terminate_backend killt fremde Sessions.
        # Auch wenn der Endpoint hinter Admin-Auth liegt, ist Defense-in-Depth
        # angemessen — ein admin-Account-Compromise soll nicht zum DB-Wipe fuehren.
        "PG_SLEEP",
        "PG_TERMINATE_BACKEND",
        "PG_CANCEL_BACKEND",
        "PG_READ_FILE",
        "PG_LS_DIR",
        "PG_RELOAD_CONF",
        "LO_IMPORT",
        "LO_EXPORT",
        "LO_PUT",
        "LO_GET",
        "DBLINK",
    ]
    import re

    for keyword in blocked:
        if re.search(rf"\b{keyword}\b", upper):
            raise ValueError(f"'{keyword}' ist in Abfragen nicht erlaubt.")

    from_count = len(re.findall(r"\bFROM\b", upper))
    select_count = len(re.findall(r"\bSELECT\b", upper))
    if select_count != 1 or from_count != 1:
        raise ValueError("Nur einfache SELECT-Abfragen mit genau einer Tabelle sind erlaubt.")

    placeholder_count = cleaned.count("{table}")
    if placeholder_count != 1:
        raise ValueError("Die Abfrage muss genau einmal '{table}' enthalten.")

    # Tabelle pruefen
    with engine.connect() as conn:
        result = conn.execute(
            text("SELECT EXISTS (SELECT 1 FROM information_schema.tables WHERE table_name = :t)"),
            {"t": table_name},
        )
        if not result.scalar():
            raise ValueError(f"Tabelle '{table_name}' nicht gefunden. Erst Daten einlesen.")

    # Nur {table} durch den sicheren Tabellennamen ersetzen
    safe_sql = cleaned.replace("{table}", f'"{table_name}"')
    if re.search(rf"\bFROM\s+(?!\"{re.escape(table_name)}\")", safe_sql, flags=re.IGNORECASE):
        raise ValueError("Die Abfrage darf nur auf die freigegebene Tabelle zugreifen.")

    # Maximal 1000 Zeilen zurueckgeben
    if "LIMIT" not in upper:
        safe_sql += " LIMIT 1000"

    # Statement-Timeout 30s: schuetzt vor Endlos-Queries (Cartesian Joins,
    # absichtlich teure WHERE-Clauses) und macht den DB-Worker robuster.
    # `SET LOCAL` benoetigt eine Transaktion -- daher engine.begin().
    with engine.begin() as conn:
        conn.execute(text("SET LOCAL statement_timeout = '30s'"))
        result = conn.execute(text(safe_sql))
        columns = list(result.keys())
        rows = [dict(zip(columns, row)) for row in result.fetchall()]

    return rows


def get_table_info(source: str) -> dict:
    """Gibt Schema-Info und Statistiken einer DataFrame-Tabelle zurueck."""
    table_name = _safe_table_name(source)

    with engine.connect() as conn:
        # Pruefen ob Tabelle existiert
        result = conn.execute(
            text("SELECT EXISTS (SELECT 1 FROM information_schema.tables WHERE table_name = :t)"),
            {"t": table_name},
        )
        if not result.scalar():
            return {"exists": False, "table_name": table_name}

        # Spalten
        cols = conn.execute(
            text("""
                SELECT column_name, data_type
                FROM information_schema.columns
                WHERE table_name = :t
                ORDER BY ordinal_position
            """),
            {"t": table_name},
        )
        columns = [{"name": r[0], "type": r[1]} for r in cols]

        # Zeilenanzahl
        count = conn.execute(text(f'SELECT COUNT(*) FROM "{table_name}"')).scalar()

    return {
        "exists": True,
        "table_name": table_name,
        "row_count": count,
        "columns": columns,
    }


def get_summary_stats(source: str) -> str:
    """
    Erzeugt eine menschenlesbare Zusammenfassung einer DataFrame-Tabelle.
    Wird als Kontext an das LLM uebergeben.
    """
    table_name = _safe_table_name(source)
    info = get_table_info(source)
    if not info.get("exists"):
        return f"Tabelle '{source}' nicht gefunden."

    lines = [
        f"Datenquelle: {source}",
        f"Tabelle: {table_name}",
        f"Zeilen: {info['row_count']}",
        f"Spalten: {', '.join(c['name'] for c in info['columns'])}",
        "",
    ]

    # Numerische Spalten: Min/Max/Avg/Sum
    numeric_cols = [c["name"] for c in info["columns"]
                    if c["type"] in ("integer", "bigint", "numeric", "double precision", "real")]

    if numeric_cols:
        lines.append("Numerische Statistiken:")
        with engine.connect() as conn:
            for col in numeric_cols[:10]:  # Max 10 Spalten
                try:
                    row = conn.execute(text(
                        f'SELECT MIN("{col}"), MAX("{col}"), AVG("{col}")::numeric(20,2), '
                        f'SUM("{col}")::numeric(20,2), COUNT("{col}") '
                        f'FROM "{table_name}" WHERE "{col}" IS NOT NULL'
                    )).fetchone()
                    if row and row[4] > 0:
                        lines.append(
                            f"  {col}: Min={row[0]}, Max={row[1]}, "
                            f"Avg={row[2]}, Summe={row[3]}, Anzahl={row[4]}"
                        )
                except Exception:
                    pass

    # Top 5 Zeilen als Beispiel
    lines.append("\nBeispieldaten (erste 5 Zeilen):")
    with engine.connect() as conn:
        rows = conn.execute(text(f'SELECT * FROM "{table_name}" LIMIT 5')).fetchall()
        col_names = [c["name"] for c in info["columns"]]
        for row in rows:
            fields = [f"{col_names[i]}: {row[i]}" for i in range(len(row))
                      if row[i] is not None and str(row[i]).strip()]
            lines.append("  " + " | ".join(fields[:6]))

    return "\n".join(lines)


def _extract_beneficiary_prompt_filters(
    prompt: str | None,
    sources: list[dict[str, Any]],
) -> tuple[str | None, str | None]:
    normalized_prompt = _normalize_search_text(prompt)
    if not normalized_prompt:
        return None, None

    bundesland = None
    known_states = sorted(
        {
            str(item.get("bundesland")).strip()
            for item in sources
            if item.get("bundesland")
        },
        key=len,
        reverse=True,
    )
    for state in known_states:
        if _normalize_search_text(state) in normalized_prompt:
            bundesland = state
            break

    fonds = None
    for candidate in ("EFRE", "ESF", "JTF", "ISF", "AMIF"):
        if candidate.lower() in normalized_prompt:
            fonds = candidate
            break

    return bundesland, fonds


# Schluesselwoerter, die einen Begueunstigten-Typ-Filter triggern.
# Pro Eintrag: (Prompt-Trigger-Tupel, Such-Substring-Tupel, Klartext-Label).
# Trigger werden in normalisiertem Prompt-Text (lowercase) gesucht — daher
# Stamm-basiert ohne Endungen, damit Umlaut-Varianten und Plural matchen.
_BENEFICIARY_TYPE_FILTERS: list[tuple[tuple[str, ...], tuple[str, ...], str]] = [
    (
        ("universit", " uni ", " tu ", " th ", "rwth"),
        ("universit", " tu ", " th ", "rwth"),
        "Universitäten",
    ),
    (
        ("hochschul", "fachhochschul", " fh "),
        ("hochschule", "fachhochschule"),
        "Hochschulen",
    ),
    (
        ("forschungseinrichtung", "forschungsinstitut", "fraunhofer", "max-planck", "max planck", "leibniz", "helmholtz", " dlr "),
        ("forschung", "institut", "fraunhofer", "max-planck", "max planck", "leibniz", "helmholtz", "dlr"),
        "Forschungseinrichtungen",
    ),
    (
        ("klinik", "krankenhaus", "spital", "uniklinik", "universitätsklinik", "universitaetsklinik"),
        ("klinik", "krankenhaus", "spital"),
        "Kliniken / Krankenhäuser",
    ),
    (
        ("kommune", "stadt ", "städt", "staedt", "gemeind", "landkreis", " kreis "),
        ("stadt ", "gemeinde ", "landkreis", "kreis "),
        "Kommunen / Gebietskörperschaften",
    ),
    (
        ("verein", "e.v.", " ev "),
        ("e.v.", "verein", " ev"),
        "Vereine",
    ),
    (
        ("gmbh",),
        ("gmbh",),
        "GmbHs",
    ),
    (
        (" ag ",),
        (" ag",),
        "Aktiengesellschaften",
    ),
    (
        ("stiftung",),
        ("stiftung",),
        "Stiftungen",
    ),
    (
        ("kmu", "mittelstand", "mittelständ", "mittelstaend"),
        ("kmu",),
        "KMU",
    ),
]


def _detect_beneficiary_name_filter(
    prompt: str | None,
) -> tuple[tuple[str, ...] | None, str | None]:
    """Erkennt Begueunstigten-Typ-Filter im Prompt.

    Gibt (Substring-Tupel, Klartext-Label) zurueck, oder (None, None).
    """
    normalized = _normalize_search_text(prompt)
    if not normalized:
        return None, None
    padded = f" {normalized} "
    for triggers, substrings, label in _BENEFICIARY_TYPE_FILTERS:
        if any(t in padded for t in triggers):
            return substrings, label
    return None, None


# Stämme, die ein grossgeschriebenes Wort als Domänen-/Auswertungsbegriff
# entlarven (auch in Komposita wie "EFRE-Volumen" oder "Fördersummen-Ranking").
# Solche Tokens sind keine Eigennamen und dürfen das Begünstigten-Namensrouting
# nicht kapern. Casefold-Vergleich (Substring).
_NON_ENTITY_STEMS = (
    "förder", "foerder", "volumen", "fördermittel", "foerdermittel",
    "zuschuss", "zuschüss", "zuschuess", "fonds", "efre", "esf", "jtf",
    "bundesland", "bundeslaender", "bundesländer", "verteilung", "aufteilung",
    "auswertung", "übersicht", "uebersicht", "ranking", "rangliste",
    "gesamtsumme", "gesamtvolumen",
)


# Wörter, die zwar grossgeschrieben sind, aber keine konkreten Eigennamen
# darstellen (Bundeslaender werden separat als Filter behandelt).
_PROPER_NOUN_STOPLIST = {
    # Frage-/Funktionswoerter
    "Welche", "Welcher", "Welches", "Wieviel", "Wieviele", "Viel", "Viele",
    "Wie", "Was", "Wer", "Wo", "Wofuer", "Wofür", "Warum",
    "Geld", "Gelder", "Mittel", "Foerderung", "Förderung", "Foerdermittel",
    "Fördermittel", "Foerderungen", "Förderungen", "Zuschuss", "Zuschüsse",
    "Bekommen", "Bekommt", "Erhalten", "Erhält", "Werden", "Wird", "Sind",
    "Hat", "Haben", "Diese", "Dieser", "Dieses", "Diesen",
    # Begueunstigten-/Foerder-Begriffe (sind keine Eigennamen!)
    "Begünstigte", "Begünstigten", "Begünstigter", "Begünstigtem",
    "Beguenstigte", "Beguenstigten", "Beguenstigter", "Beguenstigtem",
    "Träger", "Traeger", "Empfänger", "Empfaenger",
    "Vorhabens", "Vorhabentraeger", "Vorhabenträger",
    "Bundesländer", "Bundesländern", "Bundeslaender", "Bundeslaendern",
    # Mehrere/Verschiedene
    "Mehrere", "Mehreren", "Mehrerer", "Mehreres",
    "Verschiedene", "Verschiedenen", "Diverse",
    "Mehrfach", "Mehrfache", "Mehrfachen",
    # Listen-/Quellen-Begriffe
    "Listen", "Verzeichnissen", "Quellen",
    # Domaen-Begriffe rund um die Foerdersystematik (sind keine Eigennamen!)
    "Projekt", "Projekte", "Projekten", "Vorhaben",
    "Anzahl", "Summe", "Betrag", "Beträge", "Betraege",
    "Volumen", "Gesamtvolumen", "Foerderung", "Foerdervolumen",
    "Fördervolumen", "Euro", "Tausend", "Million", "Millionen",
    "Milliarde", "Milliarden", "Periode", "Foerderperiode", "Förderperiode",
    "Liste", "Verzeichnis", "Verzeichnisse", "Daten", "Datensatz",
    "Datensaetze", "Datensätze",
    # Auswertungs-/Frage-Intent-Woerter — keine Eigennamen, sondern die
    # Auswertung selbst. Ohne diese landen Saetze wie "Verteilung der
    # Foerdermittel nach Bundesland" oder "Ranking der Bundeslaender" als
    # vermeintlicher Begueunstigtenname im top_beneficiaries-Filter (0 Treffer)
    # und fallen faelschlich in den LLM-Pfad, statt state_fund_totals zu treffen.
    "Verteilung", "Verteilungen", "Verteilungs", "Mittelverteilung",
    "Aufteilung", "Aufteilungen", "Aufstellung", "Aufschlüsselung",
    "Aufschluesselung", "Übersicht", "Uebersicht", "Überblick", "Ueberblick",
    "Auswertung", "Auswertungen", "Ranking", "Rangliste", "Reihenfolge",
    "Vergleich", "Vergleiche", "Anteil", "Anteile", "Zusammenfassung",
    "Statistik", "Statistiken", "Quote", "Quoten", "Fonds", "Fördertopf",
    "Foerdertopf", "Darstellung", "Aufschluss", "Quelle",
    # Auswertungs-Dimensionen der Analytics-Kachel (sind keine Eigennamen,
    # sondern die Gruppierungsachse — sonst Hijack des top_beneficiaries-Routings)
    "Standort", "Standorte", "Standorten", "Wirtschaftszweig", "Wirtschaftszweige",
    "Wirtschaftsbereich", "Wirtschaftsbereiche", "Branche", "Branchen",
    "Sektor", "Sektoren", "Interventionsbereich", "Interventionsbereiche",
    "Interventionskategorie", "Themenfeld", "Themenfelder", "Förderbereich",
    "Foerderbereich", "Förderbereiche", "Foerderbereiche",
    # Geld-/Foerder-Synonyme (sind keine Eigennamen!)
    "Fördergelder", "Foerdergelder", "Fördersumme", "Foerdersumme",
    "Fördersummen", "Foerdersummen", "Fördervolumina", "Foerdervolumina",
    "Gesamtförderung", "Gesamtfoerderung", "Gesamtsumme", "Top",
    # Imperative / Aufforderungen (Frage-Satzanfänge der Demo-Buttons)
    "Zeig", "Zeige", "Gib", "Nenne", "Sortiere", "Gruppiere", "Zähl", "Zaehl",
    "Liste", "Vergleiche", "Berechne", "Ermittle", "Stelle", "Stell",
    "Berücksichtige", "Beruecksichtige", "Berücksichtig", "Beruecksichtig",
    "Analysiere", "Analysier", "Schlüssele", "Schluessele", "Erstelle",
    # Begueunstigten-Typen (sind schon ueber _BENEFICIARY_TYPE_FILTERS abgedeckt)
    "Universität", "Universitaet", "Universitäten", "Universitaeten",
    "Uni", "Hochschule", "Hochschulen", "Fachhochschule", "Fachhochschulen",
    "Klinik", "Kliniken", "Krankenhaus", "Krankenhäuser", "Krankenhaeuser",
    "Stiftung", "Stiftungen", "Verein", "Vereine", "Forschung",
    "Forschungseinrichtung", "Forschungseinrichtungen",
    "Stadt", "Gemeinde", "Landkreis", "Kommune", "Kommunen", "Land",
    "Bundesland", "Bundeslaender", "Bundesländer",
    # Laender / EU-Begriffe
    "Deutschland", "Österreich", "Oesterreich", "Schweiz", "Europa", "EU",
    "EFRE", "ESF", "JTF",
    # Bundeslaender (Ein-Wort-Versionen)
    "Bayern", "Sachsen", "Berlin", "Brandenburg", "Thüringen", "Thueringen",
    "Niedersachsen", "Hamburg", "Bremen", "Saarland", "Hessen",
    "Mecklenburg-Vorpommern", "Mecklenburg", "Vorpommern",
    "Nordrhein-Westfalen", "Schleswig-Holstein",
    "Rheinland-Pfalz", "Sachsen-Anhalt", "Baden-Württemberg",
    "Baden-Wuerttemberg",
}


# Header-Artefakte: erste-Zeile-Werte aus XLSX, die fälschlich als Begünstigten-
# namen interpretiert werden (Bremen ESF, Brandenburg ESF, Schleswig-Holstein
# ESF u. a. legen die Spaltennamen erst in Zeile 2-4 ab).
_HEADER_ARTEFACT_VALUES = {
    # Englisch
    "beneficiary", "beneficiary name", "beneficiarys name", "beneficiary's name",
    "benef name", "benef_name", "name of beneficiary", "operation beneficiary",
    "recipient", "recipient name", "company", "company name",
    # Deutsch
    "begunstigter", "begunstigte", "begunstigten",
    "name des begunstigten", "name des begunstigten auftragnehmers",
    "name des auftragnehmers", "name des auftragnehmers des begunstigten",
    "zuwendungsempfanger", "zuwendungsempfanger name des begunstigten",
    "name des zuwendungsempfangers",
    # Programmatische Reste / Platzhalter
    "name", "names", "name1", "operation_id", "oper_id", "nan", "none",
    "n/a", "na", "k.a.", "ka", "unknown", "unbekannt", "test", "?", "-",
}


def _is_header_artefact(name: str) -> bool:
    """Erkennt Begünstigtennamen, die in Wahrheit Header-/Platzhalter-Werte
    sind und nicht in die Aggregation einfließen sollen.
    """
    if not name:
        return True
    raw = name.strip()
    if not raw:
        return True
    # Lower + Diakritika-Approximation für robuste Stoppliste
    key = raw.casefold()
    key = (key.replace("ä", "a").replace("ö", "o").replace("ü", "u")
              .replace("ß", "ss").replace("'", "").replace("`", "")
              .replace("/", " ").replace(",", " "))
    key = re.sub(r"\s+", " ", key).strip()
    if key in _HEADER_ARTEFACT_VALUES:
        return True
    # "Beneficiary's Name" mit Sonderzeichen
    if "beneficiary" in key and ("name" in key or len(key.split()) <= 2):
        return True
    return False


# Rechtsform-Suffixe / -Synonyme, die für die Begünstigten-Konsolidierung
# entfernt werden, sodass z. B. drei Fraunhofer-Schreibweisen
# ("eingetragener Verein", "e.V.", "e. V.") in einen Bucket fallen.
_LEGAL_SUFFIX_PATTERNS = [
    r"\beingetragener\s+verein\b",
    r"\bgemein?nutziger\s+(?:eingetragener\s+)?verein\b",
    r"\be\.\s*v\.?\b",
    r"\b(?:gemein?nutzige|gemein?nutzig)\s+(?:gmbh|ag|aktiengesellschaft)\b",
    r"\bgemein?nutzige?\s*\b",
    r"\bggmbh\b", r"\bgmbh\s*&\s*co\.\s*kg\b", r"\bgmbh\b",
    r"\bag\b", r"\bkg\b", r"\bohg\b", r"\bse\b", r"\bug\b",
    r"\bmbh\b", r"\baor\b", r"\baoer\b", r"\bgesellschaft\s+mbh\b",
    r"\baktiengesellschaft\b",
    r"\bder\s+prasident(?:in)?\b", r"\bdie\s+prasident(?:in)?\b",
    r"\bder\s+rektor\b", r"\bdie\s+rektorin\b",
]


def _canonical_company_key(name: str) -> str:
    """Erzeugt einen kanonischen Vergleichsschluessel fuer Begueunstigten-
    Namen: Diakritika weg, Rechtsformsuffixe weg, Mehrfach-Whitespace weg.

    Zielsetzung: identische Einrichtungen mit unterschiedlicher Schreibweise
    fallen in denselben Bucket — z. B.
      "Fraunhofer-Gesellschaft zur Förderung der angewandten Forschung e.V."
      "Fraunhofer-Gesellschaft zur Förderung der angewandten Forschung e. V."
      "Fraunhofer-Gesellschaft zur Förderung der angewandten Forschung
       eingetragener Verein"
    werden zu demselben Key.
    """
    if not name:
        return ""
    s = name.casefold()
    s = (s.replace("ä", "a").replace("ö", "o").replace("ü", "u")
           .replace("ß", "ss"))
    s = re.sub(r"[^\w\s\-&\.]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    for pat in _LEGAL_SUFFIX_PATTERNS:
        s = re.sub(pat, " ", s)
    s = re.sub(r"[\s\-]+", " ", s).strip(" -.,&")
    return s


def _extract_prompt_proper_nouns(prompt: str | None) -> list[str]:
    """Extrahiert konkrete Eigennamen aus einem deutschen Prompt.

    Findet Wörter mit Großbuchstabe-Anfang und >= 4 Zeichen, abzüglich der
    Stoppliste. Das fängt z. B. Stadt- und Personennamen wie "Gießen",
    "Frankfurt", "Halle", "Justus-Liebig" auf.
    """
    if not prompt:
        return []
    # Trennzeichen: Whitespace, Satzzeichen außer Bindestrich (Eigennamen
    # wie "Justus-Liebig-Universität" sollen erhalten bleiben).
    candidates = re.findall(r"[A-ZÄÖÜ][a-zäöüßA-ZÄÖÜ\-]{3,}", prompt)
    return [
        c for c in candidates
        if c not in _PROPER_NOUN_STOPLIST
        # Komposita mit Domänen-Stamm (z. B. "EFRE-Volumen", "ESF-Förderung",
        # "Fördervolumen-Vergleich", "JTF-Mittel") sind keine Eigennamen, sondern
        # Auswertungsbegriffe — sonst kapern sie das top_beneficiaries-Routing.
        and not any(stem in c.casefold() for stem in _NON_ENTITY_STEMS)
    ]


# Heuristisches Mapping Stadt → Bundesland fuer den Fallback-Pfad.
# Wenn der Prompt eine dieser Staedte enthaelt, kann der LLM-Kontext auf
# Begueunstigte aus diesem Bundesland eingeschraenkt werden — auch wenn der
# Begueunstigte selbst die Stadt nicht im Namen fuehrt (z. B. THM = FH
# Gießen, Sitz Gießen, aber Name "Mittelhessen").
_CITY_TO_STATE: dict[str, str] = {
    # Hessen
    "frankfurt am main": "Hessen",
    "frankfurt": "Hessen",  # weicher default; bei "Frankfurt Oder" siehe Brandenburg
    "wiesbaden": "Hessen", "kassel": "Hessen", "darmstadt": "Hessen",
    "marburg": "Hessen", "gießen": "Hessen", "giessen": "Hessen",
    "fulda": "Hessen", "hanau": "Hessen", "offenbach": "Hessen",
    # Baden-Württemberg
    "stuttgart": "Baden-Württemberg", "karlsruhe": "Baden-Württemberg",
    "freiburg": "Baden-Württemberg", "mannheim": "Baden-Württemberg",
    "heidelberg": "Baden-Württemberg", "tübingen": "Baden-Württemberg",
    "tuebingen": "Baden-Württemberg", "ulm": "Baden-Württemberg",
    "konstanz": "Baden-Württemberg",
    # Bayern
    "münchen": "Bayern", "muenchen": "Bayern", "nürnberg": "Bayern",
    "nuernberg": "Bayern", "augsburg": "Bayern", "regensburg": "Bayern",
    "würzburg": "Bayern", "wuerzburg": "Bayern", "erlangen": "Bayern",
    "bayreuth": "Bayern", "passau": "Bayern", "bamberg": "Bayern",
    # Nordrhein-Westfalen
    "köln": "Nordrhein-Westfalen", "koeln": "Nordrhein-Westfalen",
    "düsseldorf": "Nordrhein-Westfalen", "duesseldorf": "Nordrhein-Westfalen",
    "dortmund": "Nordrhein-Westfalen", "essen": "Nordrhein-Westfalen",
    "bochum": "Nordrhein-Westfalen", "aachen": "Nordrhein-Westfalen",
    "münster": "Nordrhein-Westfalen", "muenster": "Nordrhein-Westfalen",
    "bonn": "Nordrhein-Westfalen", "siegen": "Nordrhein-Westfalen",
    "paderborn": "Nordrhein-Westfalen", "wuppertal": "Nordrhein-Westfalen",
    "bielefeld": "Nordrhein-Westfalen",
    # Niedersachsen
    "hannover": "Niedersachsen", "göttingen": "Niedersachsen",
    "goettingen": "Niedersachsen", "braunschweig": "Niedersachsen",
    "oldenburg": "Niedersachsen", "osnabrück": "Niedersachsen",
    "osnabrueck": "Niedersachsen", "lüneburg": "Niedersachsen",
    "lueneburg": "Niedersachsen",
    # Sachsen
    "dresden": "Sachsen", "leipzig": "Sachsen", "chemnitz": "Sachsen",
    "freiberg": "Sachsen",
    # Sachsen-Anhalt
    "magdeburg": "Sachsen-Anhalt", "halle": "Sachsen-Anhalt",
    # Thüringen
    "jena": "Thüringen", "erfurt": "Thüringen", "weimar": "Thüringen",
    "ilmenau": "Thüringen",
    # Brandenburg
    "potsdam": "Brandenburg", "cottbus": "Brandenburg",
    "frankfurt (oder)": "Brandenburg", "frankfurt oder": "Brandenburg",
    # Mecklenburg-Vorpommern
    "rostock": "Mecklenburg-Vorpommern", "greifswald": "Mecklenburg-Vorpommern",
    "wismar": "Mecklenburg-Vorpommern",
    # Schleswig-Holstein
    "kiel": "Schleswig-Holstein", "lübeck": "Schleswig-Holstein",
    "luebeck": "Schleswig-Holstein", "flensburg": "Schleswig-Holstein",
    # Stadtstaaten
    "berlin": "Berlin", "hamburg": "Hamburg", "bremen": "Bremen",
    # Rheinland-Pfalz
    "mainz": "Rheinland-Pfalz", "trier": "Rheinland-Pfalz",
    "kaiserslautern": "Rheinland-Pfalz", "koblenz": "Rheinland-Pfalz",
    "landau": "Rheinland-Pfalz",
    # Saarland
    "saarbrücken": "Saarland", "saarbruecken": "Saarland",
}


def _state_from_proper_nouns(proper_nouns: list[str]) -> str | None:
    for noun in proper_nouns:
        key = noun.casefold().replace("ß", "ß")
        if key in _CITY_TO_STATE:
            return _CITY_TO_STATE[key]
    return None


def _select_beneficiary_analysis_mode(prompt: str | None) -> str:
    matched_mode = _match_beneficiary_analysis_mode(prompt)
    return matched_mode or "top_beneficiaries"


def _match_beneficiary_analysis_mode(prompt: str | None) -> str | None:
    normalized_prompt = _normalize_search_text(prompt)
    if not normalized_prompt:
        return None

    # Begünstigte, die in MEHREREN Bundesländern/Listen auftauchen.
    # WICHTIG: vor allen anderen Triggern, sonst greift z. B. "Bundesländern"
    # als Eigenname und faellt auf top_beneficiaries.
    if any(
        phrase in normalized_prompt
        for phrase in (
            "mehrere bundeslaender",
            "mehreren bundeslaendern",
            "mehrere bundesländer",
            "mehreren bundesländern",
            "verschiedene bundeslaender",
            "verschiedenen bundeslaendern",
            "verschiedene bundesländer",
            "verschiedenen bundesländern",
            "mehr als einem bundesland",
            "mehr als ein bundesland",
            "mehr als einem bundesländern",
            "in mehr als einem land",
            "mehrere listen",
            "in mehreren listen",
            "mehrere verzeichnisse",
            "in mehreren verzeichnissen",
            "mehreren quellen",
            "verschiedenen quellen",
            "bundeslandübergreifend",
            "bundeslanduebergreifend",
            "länderübergreifend",
            "laenderuebergreifend",
        )
    ):
        return "multi_state_beneficiaries"

    # Begueunstigten-Typ-Filter (Universitaet, Hochschule, Klinik, GmbH, ...) ist
    # eine explizite Entitaets-Absicht und gewinnt vor allem anderen.
    if _detect_beneficiary_name_filter(prompt)[0] is not None:
        return "top_beneficiaries"

    # STARKE Aggregations-Signale (eindeutige Gruppierungs-Achsen) werden VOR der
    # Eigennamen-Heuristik geprueft. Sonst kapert ein grossgeschriebenes
    # Dimensions-/Satzanfangs-Wort ("Verteilung", "Standorte", "Wirtschaftszweige",
    # "Muster", "Stelle") das Routing und die Auswertung faellt faelschlich in den
    # LLM-Pfad, statt die deterministische Aggregation zu liefern.

    # ANZAHL der Vorhaben je Bundesland (zählen, nicht Volumen). Vor dem
    # state_fund_totals-Block, weil "bundesland" dort sonst zuerst greift und
    # statt der Anzahl das Fördervolumen liefert.
    if any(
        phrase in normalized_prompt
        for phrase in (
            "wie viele vorhaben", "wieviele vorhaben", "wie viele projekte",
            "wieviele projekte", "anzahl der vorhaben", "anzahl vorhaben",
            "anzahl der projekte", "anzahl projekte", "zahl der vorhaben",
            "wie viele gefoerderte", "wie viele geförderte",
        )
    ) and any(
        phrase in normalized_prompt
        for phrase in (
            "bundesland", "bundesländer", "bundeslaender", "land", "länder",
            "laender", "region", "regionen",
        )
    ):
        return "region_project_counts"

    if any(
        phrase in normalized_prompt
        for phrase in (
            "mehrere vorhaben",
            "mehrfach",
            "mehrere projekte",
            "mehr als ein vorhaben",
            "mehr als ein projekt",
            "wiederholt gefoerdert",
        )
    ):
        return "repeat_beneficiaries"

    if any(
        phrase in normalized_prompt
        for phrase in (
            "bundesland",
            "bundesländer",
            "bundeslaender",
            "verteilung",
            "aufteilung",
            "pro fonds",
            "nach fonds",
            "je fonds",
            "pro land",
            "nach land",
        )
    ):
        return "state_fund_totals"

    if any(
        phrase in normalized_prompt
        for phrase in (
            "kommune",
            "kommunen",
            "standort",
            "standorte",
            "stadt",
            "städte",
            "staedte",
            "orte",
            "landkreis",
            "landkreise",
        )
    ):
        return "top_locations"

    if any(
        phrase in normalized_prompt
        for phrase in (
            "wirtschaftszweig",
            "wirtschaftszweige",
            "wirtschaftstaetigkeit",
            "wirtschaftstätigkeit",
            "wirtschaftsbereich",
            "branche",
            "branchen",
            "sektor",
            "sektoren",
            "nace",
            "interventionsbereich",
            "interventionsbereiche",
            "interventionskategorie",
            "art der intervention",
            "intervention field",
            "intervention",
            "foerderbereich",
            "förderbereich",
            "foerderbereiche",
            "förderbereiche",
            "themenfeld",
            "themenfelder",
            "themengebiet",
            "themengebiete",
        )
    ):
        return "top_sectors"

    # Konkrete Eigennamen (Firmen-, Personen-, Stadtname) ⇒ gefilterte Top-Liste.
    # NACH den starken Aggregations-Signalen (reine Dimensionswoerter zaehlen dann
    # nicht mehr als Eigenname), aber VOR den schwachen Fondsnamen.
    if _extract_prompt_proper_nouns(prompt):
        return "top_beneficiaries"

    # SCHWACHE Signale: blosse Fondsnamen ohne Gruppierungsachse. Erst hier, damit
    # z. B. "Wie viel EFRE-Geld hat Siemens bekommen?" beim Eigennamen bleibt.
    if any(phrase in normalized_prompt for phrase in ("fonds", "efre", "esf", "jtf")):
        return "state_fund_totals"

    if any(
        phrase in normalized_prompt
        for phrase in (
            "groesste beguenstigte",
            "größte begünstigte",
            "groessten beguenstigten",
            "größten begünstigten",
            "top beguenstigte",
            "top-beguenstigte",
            "top begünstigte",
            "beguenstigte",
            "begünstigte",
            "traeger",
            "träger",
        )
    ):
        return "top_beneficiaries"

    return None


def get_beneficiary_llm_context(
    prompt: str | None = None,
    max_entries_per_source: int = 3,
    country_code: str | None = None,
) -> str:
    sources = get_beneficiary_sources(country_code=country_code)
    if not sources:
        return ""

    bundesland, fonds = _extract_beneficiary_prompt_filters(prompt, sources)
    mode = _select_beneficiary_analysis_mode(prompt)
    type_substrings, _name_filter_label = _detect_beneficiary_name_filter(prompt)
    proper_nouns = _extract_prompt_proper_nouns(prompt)
    name_substrings: list[tuple[str, ...]] = []
    if type_substrings:
        name_substrings.append(type_substrings)
    for noun in proper_nouns:
        name_substrings.append((noun,))
    limit = max(4, min(max_entries_per_source * 3, 12))
    analysis = analyze_beneficiary_records(
        mode=mode,
        bundesland=bundesland,
        fonds=fonds,
        limit=limit,
        country_code=country_code,
        name_substrings=name_substrings or None,
    )
    detail_query = " ".join(proper_nouns).strip()
    if not detail_query and prompt and any(
        phrase in _normalize_search_text(prompt)
        for phrase in ("vorhaben von", "projekte von", "foerderung fuer", "förderung für", "traeger ", "träger ")
    ):
        detail_query = prompt.strip()
    detail_results: dict[str, Any] | None = None
    if detail_query:
        try:
            detail_results = search_beneficiary_records(
                query=detail_query,
                scope="all",
                bundesland=bundesland,
                fonds=fonds,
                limit=12,
                company_limit=6,
                country_code=country_code,
            )
        except Exception:
            log.exception("Detailkontext fuer Begueunstigtenfrage fehlgeschlagen.")
            detail_results = None

    # Wenn die strenge Filterung leer ist, dem LLM zusaetzlich die
    # *naechstgelegenen* Kandidaten liefern. So kann das Modell
    # historische Namen aufloesen (FH Gießen → Technische Hochschule
    # Mittelhessen), Eigenschreibweisen und Mehrfach-Ortsnamen klaeren.
    fallback_items: list[dict[str, Any]] = []
    fallback_state = bundesland or _state_from_proper_nouns(proper_nouns)
    if not analysis["items"] and (proper_nouns or type_substrings):
        # Zuerst nur mit Typ-Filter (ohne Eigenname) erneut analysieren —
        # bevorzugt im heuristisch erkannten Bundesland (z. B. "Gießen" → Hessen)
        if type_substrings:
            broad = analyze_beneficiary_records(
                mode="top_beneficiaries",
                bundesland=fallback_state,
                fonds=fonds,
                limit=limit,
                country_code=country_code,
                name_substrings=[type_substrings],
            )
            fallback_items.extend(broad.get("items") or [])
            # Wenn der bundeslandgefilterte Pass leer war, ohne Bundesland-Filter
            if not fallback_items and fallback_state:
                broad2 = analyze_beneficiary_records(
                    mode="top_beneficiaries",
                    bundesland=None,
                    fonds=fonds,
                    limit=limit,
                    country_code=country_code,
                    name_substrings=[type_substrings],
                )
                fallback_items.extend(broad2.get("items") or [])
        # Plus rapidfuzz-Topmatches gegen den Original-Prompt
        try:
            from rapidfuzz import fuzz, process
            with engine.connect() as conn:
                # Alle Begueunstigtennamen einsammeln (nur die paar tausend
                # eindeutigen Namen, nicht die volle Tabelle pro Vorhaben)
                names: list[tuple[str, str]] = []
                for source_info in (
                    [s for s in sources if (not bundesland or (s.get("bundesland") or "") == bundesland)]
                ):
                    info = get_table_info(source_info["source"])
                    if not info.get("exists"):
                        continue
                    name_col = next(
                        (c["name"] for c in info["columns"]
                         if any(p in c["name"].lower() for p in ("auftragnehmer", "beguenstig", "begünstig"))),
                        None,
                    )
                    if not name_col:
                        continue
                    rows = conn.execute(
                        text(f'SELECT DISTINCT {_quote_ident(name_col)} FROM "{_safe_table_name(source_info["source"])}" LIMIT 2000')
                    ).fetchall()
                    for row in rows:
                        if row[0]:
                            names.append((str(row[0]).strip(), source_info.get("bundesland") or ""))
            query_text = " ".join(proper_nouns) or " ".join(type_substrings or [])
            if names and query_text:
                top_matches = process.extract(
                    query_text,
                    [n[0] for n in names],
                    scorer=fuzz.token_set_ratio,
                    limit=8,
                    score_cutoff=55,
                )
                seen = {item.get("label") for item in fallback_items}
                for matched_name, score, idx in top_matches:
                    if matched_name in seen:
                        continue
                    seen.add(matched_name)
                    fallback_items.append({
                        "label": matched_name,
                        "value_label": "",
                        "value": None,
                        "project_count": None,
                        "sublabel": f"ähnlicher Name (Score {int(score)}) · {names[idx][1]}",
                        "rank": len(fallback_items) + 1,
                    })
        except Exception:
            log.exception("rapidfuzz-Fallback fuer Begueunstigte fehlgeschlagen.")

    summary = analysis["summary"]
    filters = analysis["filters"]
    items = analysis["items"]

    parts = [
        "Strukturierte Auswertung der geladenen Beguenstigtenverzeichnisse.",
        f"Fokus: {analysis['title']}",
        (
            "Abdeckung: "
            f"{_de_int(summary['sources_considered'])} Quelle(n), "
            f"{_de_int(summary['records_scanned'])} Datensaetze, "
            f"{summary['total_volume_label']} Gesamtvolumen"
        ),
    ]

    active_filters = []
    if country_code:
        active_filters.append(f"Land={get_country_name(country_code) or country_code}")
    if filters.get("bundesland"):
        active_filters.append(f"Bundesland={filters['bundesland']}")
    if filters.get("fonds"):
        active_filters.append(f"Fonds={filters['fonds']}")
    if active_filters:
        parts.append("Filter: " + ", ".join(active_filters))

    if detail_results and detail_results.get("records"):
        detail_summary = detail_results.get("summary") or {}
        parts.append(
            "Konkrete Treffer aus der strukturierten Suche "
            f"({_de_int(detail_summary.get('matches', 0))} Treffer, "
            f"{_de_int(detail_summary.get('sources_considered', 0))} Quelle(n), "
            f"Suchbegriff='{detail_query}'):"
        )
        for row in (detail_results.get("records") or [])[:10]:
            detail_parts = [
                row.get("company_name") or "",
                row.get("project_name") or "",
                row.get("kosten_label") or "",
                row.get("location") or "",
                " / ".join(str(part) for part in (row.get("bundesland"), row.get("fonds"), row.get("periode")) if part),
                f"Quelle={row.get('source')}" if row.get("source") else "",
            ]
            parts.append("- " + " | ".join(str(part) for part in detail_parts if part))
        parts.append(
            "Nutze diese konkreten Treffer fuer Fragen nach einzelnen Traegern, "
            "Vorhaben, Orten oder Aktenzeichen. Nenne keine weiteren Einzelvorhaben."
        )

    if not items:
        if fallback_items:
            parts.append(
                "Direkter Substring-Treffer schlaegt fehl. Wahrscheinliche Kandidaten "
                "aus den Daten (zur semantischen Aufloesung — z. B. historische Namen "
                "wie 'FH Gießen' = heute 'Technische Hochschule Mittelhessen'):"
            )
            for item in fallback_items[:10]:
                detail_parts = [
                    f"{item['rank']}. {item['label']}",
                    item.get("value_label") or "",
                ]
                if item.get("project_count"):
                    detail_parts.append(f"{_de_int(item['project_count'])} Vorhaben")
                if item.get("sublabel"):
                    detail_parts.append(str(item["sublabel"]))
                parts.append("- " + " | ".join(str(p) for p in detail_parts if p))
            parts.append(
                "Pruefe, ob einer dieser Eintraege die gemeinte Einrichtung ist "
                "(z. B. nach Standort, Rechtsform oder historischem Namen). "
                "Wenn ja, beantworte die Frage anhand des passenden Eintrags. "
                "Wenn keiner passt, sage das ehrlich."
            )
            return "\n".join(parts)
        parts.append("Keine passenden Auswertungsdaten fuer diese Fragestellung vorhanden.")
        parts.append("Wenn die Frage darueber hinausgeht, weise auf die fehlende Datengrundlage hin.")
        return "\n".join(parts)

    parts.append("Rangliste:")
    for item in items:
        detail_parts = [
            f"{item['rank']}. {item['label']}",
            item.get("value_label") or _format_eur(item.get("value")),
        ]
        if item.get("project_count"):
            detail_parts.append(f"{_de_int(item['project_count'])} Vorhaben")
        if item.get("sublabel"):
            detail_parts.append(str(item["sublabel"]))
        parts.append("- " + " | ".join(str(part) for part in detail_parts if part))

    parts.append(
        "Nutze nur diese Daten. Wenn die Nutzerfrage weitere Einzelwerte, Vollstaendigkeit "
        "oder Detailbelege verlangt, benenne die Grenze klar."
    )
    return "\n".join(parts)


def _build_state_fund_totals_answer(
    prompt: str,
    country_code: str | None,
    bundesland: str | None,
    fonds: str | None,
) -> str:
    """Lesbare Rangliste 'Fördervolumen pro Bundesland'.

    Re-aggregiert die (Bundesland × Fonds)-Buckets aus
    ``analyze_beneficiary_records`` auf Bundesland-Ebene (Fonds zusammengefasst,
    aber mit Aufschlüsselung), zeigt ALLE Länder statt nur Top-5 und trennt
    Einträge ohne Länderzuordnung (bundesweite Programme wie AMIF/ISF, fehlende
    Bundesland-Spalte) als Fußnote ab.
    """
    # "Bundesland" ist ein nationaler Begriff und nur DE führt eine
    # Bundesland-Spalte. Ohne explizite Landauswahl daher Deutschland —
    # sonst mischt sich z. B. Österreich als "Unbekannt" in die Rangliste.
    eff_cc = _default_de_country(prompt, country_code)

    analysis = analyze_beneficiary_records(
        mode="state_fund_totals",
        bundesland=bundesland,
        fonds=fonds,
        limit=100,
        country_code=eff_cc,
    )
    items = analysis["items"]
    summary = analysis["summary"]
    country_label = get_country_name(eff_cc) or eff_cc

    if not items:
        return (
            f"Für {country_label} liegen in den aktuell geladenen "
            "Begünstigtenverzeichnissen keine auswertbaren Fördervolumina vor."
        )

    # Re-Aggregation auf Bundesland-Ebene; Einträge ohne Land separat sammeln.
    per_state: dict[str, dict[str, Any]] = {}
    no_state: dict[str, Any] = {"total": 0.0, "pc": 0, "fonds": {}}
    for it in items:
        bl = (it.get("bundesland") or "").strip()
        target = (
            no_state if (not bl or bl == "Unbekannt")
            else per_state.setdefault(bl, {"total": 0.0, "pc": 0, "fonds": {}})
        )
        val = float(it.get("value") or 0.0)
        target["total"] += val
        target["pc"] += int(it.get("project_count") or 0)
        fonds_name = (it.get("fonds") or "").strip() or "—"
        target["fonds"][fonds_name] = target["fonds"].get(fonds_name, 0.0) + val

    # Header-Bausteine: Fonds-Geltungsbereich und (falls einheitlich) Periode.
    if fonds:
        scope = f"nur {fonds}"
    else:
        fonds_present = sorted({
            f for d in per_state.values() for f, v in d["fonds"].items()
            if v > 0 and f != "—"
        })
        scope = " + ".join(fonds_present) if fonds_present else "alle Fonds"
    periods = {s.get("periode") for s in get_beneficiary_sources(country_code=eff_cc) if s.get("periode")}
    period_suffix = f" · Förderperiode {next(iter(periods))}" if len(periods) == 1 else ""

    def _fonds_breakdown(fonds_map: dict[str, float], sep: str = " · ") -> str:
        parts = [
            f"{f} {_format_eur_compact(v)}"
            for f, v in sorted(fonds_map.items(), key=lambda kv: -kv[1])
            if v > 0 and f != "—"
        ]
        return sep.join(parts[:3])

    lines: list[str] = []
    if per_state:
        lines.append(
            f"Fördervolumen pro Bundesland ({country_label} · {scope}{period_suffix}):"
        )
        ranked = sorted(per_state.items(), key=lambda kv: -kv[1]["total"])
        for rank, (bl, d) in enumerate(ranked, start=1):
            breakdown = _fonds_breakdown(d["fonds"])
            # Aufschlüsselung nur zeigen, wenn mehr als ein Fonds beiträgt.
            suffix = f" ({breakdown})" if len([1 for f, v in d["fonds"].items() if v > 0 and f != "—"]) > 1 else ""
            lines.append(
                f"{rank}. {bl} — {_format_eur_compact(d['total'])} "
                f"· {_de_int(d['pc'])} Vorhaben{suffix}"
            )
        if no_state["total"] > 0:
            fed = _fonds_breakdown(no_state["fonds"], sep=", ")
            fed_hint = f" ({fed})" if fed else ""
            lines.append(
                f"Hinzu kommen {_format_eur_compact(no_state['total'])} aus "
                f"bundesweiten Programmen ohne Länderzuordnung{fed_hint}."
            )
        grundlage = (
            f"Datengrundlage: {country_label}, "
            f"{'1 Bundesland' if len(per_state) == 1 else f'{_de_int(len(per_state))} Bundesländer'}, "
            f"{_de_int(summary['records_scanned'])} Datensätze, "
            f"{summary['total_volume_label']} Gesamtvolumen."
        )
    else:
        # Land ohne Bundesland-Spalte (z. B. Österreich) → Fonds-Summen.
        lines.append(
            f"{country_label} weist in den geladenen Transparenzlisten keine "
            "Bundesland-Spalte aus — eine länderscharfe Aufteilung ist hier nicht "
            "möglich. Fonds-Summen:"
        )
        for rank, (f, v) in enumerate(
            sorted(no_state["fonds"].items(), key=lambda kv: -kv[1]), start=1
        ):
            if v > 0 and f != "—":
                lines.append(f"{rank}. {f} — {_format_eur_compact(v)}")
        grundlage = (
            f"Datengrundlage: {country_label}, "
            f"{_de_int(summary['records_scanned'])} Datensätze, "
            f"{summary['total_volume_label']} Gesamtvolumen."
        )

    lines.append(grundlage)
    return "\n".join(lines)


# Werte, die in Kategorie-Spalten als Header-/Platzhalter-Reste auftauchen und
# keine echte Kategorie sind. Casefold-Vergleich nach Whitespace-Normalisierung.
_CATEGORY_JUNK = {
    "", "-", "–", "—", ".", "nan", "none", "na", "n/a", "k.a.", "ka", "kein",
    "keine", "keine angabe", "nicht zutreffend", "not applicable",
    "spezifisches ziel", "specific objective", "type of intervention",
    "type of intervation", "art der intervention",
    "art der intervention für das vorhaben", "interventionsbereich",
    "intervention field", "intervention", "categorisation",
    "category of intervention", "economic activity", "priority",
    "prg_priority", "country", "efre", "esf", "jtf", "amif", "isf",
    # geleakte Summen-/Total-Zeilen
    "gesamt", "summe", "gesamt - summe", "gesamtsumme", "summe gesamt",
    "total", "insgesamt", "zwischensumme",
}

# Werte, die in Standort-Spalten kein verwertbarer Ort sind.
_LOCATION_JUNK = {
    "", "-", "–", "nan", "none", "na", "k.a.", "ka", "ort", "standort",
    "region", "nuts", "nuts2", "nuts3", "plz", "stadt", "gemeinde", "country",
    "unbekannt", "coordinates", "koordinaten", "deutschland", "österreich",
    "oesterreich",
}

# Flächen-Bundesländer sind kein „Standort"; die Stadtstaaten Berlin/Hamburg/
# Bremen dagegen schon und bleiben erhalten.
_NON_CITY_STATES = {
    "baden-württemberg", "baden-wuerttemberg", "bayern", "sachsen",
    "sachsen-anhalt", "thüringen", "thueringen", "brandenburg",
    "niedersachsen", "hessen", "nordrhein-westfalen", "rheinland-pfalz",
    "schleswig-holstein", "mecklenburg-vorpommern", "saarland",
}


def _classify_category_taxonomy(column_name: str | None) -> str | None:
    """Ordnet eine erkannte Kategorie-Spalte einem der drei Klassifikations-
    systeme zu — anhand des Spaltennamens, damit es nach jedem Harvest stabil
    bleibt (keine Bindung an konkrete Quellen).
    """
    if not column_name:
        return None
    c = column_name.casefold()
    # NACE zuerst: kombinierte Spalten ("wirtschaftstätigkeit … art der
    # intervention …") tragen NACE-Codes.
    if any(k in c for k in ("wirtschaftstätigkeit", "wirtschaftstaetigkeit",
                            "wirtschaftszweig", "economic activity", "nace")):
        return "nace"
    if any(k in c for k in ("spezifisch", "specific objective", "prg_priority",
                            "priorität", "prioritaet", "prg_priority")):
        return "objective"
    if any(k in c for k in ("intervention", "interventionsbereich",
                            "interventionskategorie", "categorisation")):
        return "intervention"
    return None


def _is_category_artefact(value: Any) -> bool:
    norm = re.sub(r"\s+", " ", str(value or "").replace("\n", " ")).strip()
    low = norm.casefold()
    if low in _CATEGORY_JUNK:
        return True
    # Geleakte Header-Zeile (z. B. Berlin EFRE)
    if "categories of intervention for the operation" in low:
        return True
    return False


def _category_bucket_key(value: str) -> tuple[str, str]:
    """Liefert (Gruppenschlüssel, Anzeige-Code-Hinweis) für eine Kategorie.

    Führende Nummern-Codes ("004 - …", "1.1 …", "136") werden normalisiert, sodass
    dieselbe Kategorie über Listen mit/ohne Bezeichnung in EINEN Bucket fällt.
    """
    v = re.sub(r"\s+", " ", value).strip()
    m = re.match(r"^(\d{1,4}(?:\.\d+)*)\s*[-–:.)]?\s*", v)
    if m:
        code = m.group(1)
        # Reine Ganzzahl-Codes ohne Punkt: führende Nullen ignorieren ("004"=="4").
        norm = code if "." in code else (code.lstrip("0") or "0")
        return f"code:{norm}", code
    return f"text:{v.casefold()}", ""


def _scan_beneficiary_records(
    country_code: str | None,
    bundesland: str | None,
    fonds: str | None,
):
    """Generator über alle passenden Begünstigten-Datensätze.

    Liest Quellen und Spalten bei JEDEM Aufruf live aus der DB — nach einem
    erneuten Harvest spiegelt die Auswertung sofort den neuen Stand. Yieldet je
    Datensatz die Roh-Kategorie (+ erkannte Spalte/Taxonomie), den Roh-Standort,
    die geparsten Kosten sowie Fonds/Bundesland aus den Quellen-Metadaten.
    """
    from services.geocoding_service import detect_columns

    sources = [
        s for s in get_beneficiary_sources(country_code=country_code)
        if (not bundesland or (s.get("bundesland") or "") == bundesland)
        and (not fonds or (s.get("fonds") or "") == fonds)
    ]
    for s in sources:
        src = s["source"]
        table = _safe_table_name(src)
        cols = detect_columns(src)
        cat_col = cols.get("sz")
        loc_col = cols.get("standort") or cols.get("ort") or cols.get("landkreis")
        cost_col = cols.get("kosten")
        taxonomy = _classify_category_taxonomy(cat_col)
        selected = [
            (alias, col) for alias, col in (
                ("category", cat_col), ("location", loc_col), ("kosten", cost_col),
            ) if col
        ]
        if not selected:
            continue
        sql = ", ".join(f"{_quote_ident(col)} AS {alias}" for alias, col in selected)
        try:
            with engine.connect() as conn:
                rows = conn.execute(text(f'SELECT {sql} FROM "{table}"')).fetchall()
        except Exception:
            log.exception("Scan der Quelle %s fehlgeschlagen.", src)
            continue
        for row in rows:
            entry = dict(row._mapping)
            yield {
                "category": entry.get("category"),
                "taxonomy": taxonomy,
                "location": entry.get("location"),
                "kosten": _parse_numeric(entry.get("kosten")),
                "fonds": s.get("fonds"),
                "bundesland": s.get("bundesland"),
            }


def _default_de_country(prompt: str, country_code: str | None) -> str | None:
    """Ohne explizites Land Deutschland annehmen, außer der Prompt nennt
    ausdrücklich Österreich (verhindert AT/DE-Code-Mischung in Aggregaten)."""
    if country_code:
        return country_code
    norm = _normalize_search_text(prompt) or ""
    if any(tok in norm for tok in ("österreich", "oesterreich", "austria")):
        return "AT"
    return "DE"


def _build_top_sectors_answer(
    prompt: str,
    country_code: str | None,
    bundesland: str | None,
    fonds: str | None,
) -> str:
    """Geförderte Bereiche — getrennt nach den drei Klassifikationssystemen
    (Interventionsbereich · NACE-Wirtschaftszweig · Spezifisches Ziel), weil die
    Quelllisten diese uneinheitlich führen und ein Mischen fachlich falsch wäre.
    """
    eff_cc = _default_de_country(prompt, country_code)
    country_label = get_country_name(eff_cc) or eff_cc

    # taxonomy -> bucket_key -> {value, count, label, fonds:set, bl:set}
    grouped: dict[str, dict[str, dict[str, Any]]] = {
        "intervention": {}, "nace": {}, "objective": {},
    }
    scanned = 0
    classified = 0
    for rec in _scan_beneficiary_records(eff_cc, bundesland, fonds):
        scanned += 1
        tax = rec["taxonomy"]
        if tax not in grouped:
            continue
        raw = str(rec["category"] or "").strip()
        if not raw or _is_category_artefact(raw):
            continue
        classified += 1
        key, _code = _category_bucket_key(raw)
        bucket = grouped[tax].setdefault(key, {
            "value": 0.0, "count": 0, "label": "", "fonds": set(), "bl": set(),
        })
        if rec["kosten"] is not None:
            bucket["value"] += float(rec["kosten"])
        bucket["count"] += 1
        # Längste/aussagekräftigste Original-Schreibweise als Anzeige-Label.
        if len(raw) > len(bucket["label"]):
            bucket["label"] = raw
        if rec["fonds"]:
            bucket["fonds"].add(rec["fonds"])
        if rec["bundesland"]:
            bucket["bl"].add(rec["bundesland"])

    if not classified:
        return (
            f"Für {country_label} liegen in den geladenen Begünstigtenverzeichnissen "
            "keine auswertbaren Kategorieangaben (Interventionsbereich, "
            "Wirtschaftszweig oder Spezifisches Ziel) vor."
        )

    section_titles = {
        "intervention": "Interventionsbereiche (Art der Intervention · Anhang I CPR)",
        "nace": "Wirtschaftszweige (NACE)",
        "objective": "Spezifische Ziele / Prioritäten",
    }
    lines = [f"Geförderte Bereiche ({country_label}), getrennt nach Klassifikation:"]
    for tax in ("intervention", "nace", "objective"):
        buckets = grouped[tax]
        if not buckets:
            continue
        ranked = sorted(
            buckets.values(), key=lambda b: (-float(b["value"] or 0.0), -b["count"])
        )[:6]
        lines.append("")
        lines.append(f"{section_titles[tax]}:")
        for rank, b in enumerate(ranked, start=1):
            label = b["label"]
            # Reiner Code ohne Bezeichnung lesbarer machen ("1" → "Code 1").
            if not re.search(r"[A-Za-zÄÖÜäöü]", label):
                label = f"Code {label}"
            label = label if len(label) <= 90 else label[:87].rstrip() + "…"
            val = _format_eur_compact(b["value"])
            lines.append(f"{rank}. {label} — {val} · {_de_int(b['count'])} Vorhaben")

    lines.append("")
    lines.append(
        f"Datengrundlage: {country_label}, {_de_int(classified)} von "
        f"{_de_int(scanned)} Vorhaben mit auswertbarer Kategorieangabe."
    )
    return "\n".join(lines)


def _clean_location_value(value: Any) -> str | None:
    """Normalisiert einen Standort-Rohwert auf einen lesbaren Ortsnamen oder
    verwirft ihn (NUTS-Code, reine PLZ, Header-Rest, Flächen-Bundesland)."""
    v = re.sub(r"\s+", " ", str(value or "")).strip()
    if not v or v.casefold() in _LOCATION_JUNK:
        return None
    low = v.casefold()
    if "standortindikator" in low or "geolokalisier" in low:
        return None
    # Vollständige Adresse "…, 12345 Stadt" → Stadt
    m = re.search(r"\b\d{5}\s+([A-Za-zÄÖÜäöüß][\wÄÖÜäöüß .\-]+)$", v)
    if m:
        v = m.group(1).strip()
    else:
        # "12345 - Stadt" / "12345 Stadt" → Stadt
        v = re.sub(r"^\d{4,5}\s*[-–]?\s*", "", v).strip()
    # Mehrfach-Ort ("Ulm, Tübingen, …") → erster Ort
    v = v.split(",")[0].strip(" -–")
    # angehängten NUTS-Code entfernen ("Deutschland (DE80)" → "Deutschland")
    v = re.sub(r"\s*\(\s*[A-Z]{2}[0-9A-Z]{1,4}\s*\)\s*$", "", v).strip()
    if not v or v.casefold() in _LOCATION_JUNK:
        return None
    # Reiner NUTS-Code (DE, DE80, DE40C, …) oder reine PLZ
    if re.fullmatch(r"[A-Z]{2}[0-9A-Z]{0,4}", v) or re.fullmatch(r"\d{3,5}", v):
        return None
    if v.casefold() in _NON_CITY_STATES:
        return None
    if not re.search(r"[A-Za-zÄÖÜäöü]", v):
        return None
    return v


def _build_top_locations_answer(
    prompt: str,
    country_code: str | None,
    bundesland: str | None,
    fonds: str | None,
) -> str:
    """Standorte mit dem höchsten Fördervolumen — auf eine konsistente Ortsebene
    normalisiert (NUTS-Codes, reine PLZ, Adress-Präfixe und Flächen-Bundesländer
    werden bereinigt) mit ehrlicher Abdeckungsangabe."""
    eff_cc = _default_de_country(prompt, country_code)
    country_label = get_country_name(eff_cc) or eff_cc

    grouped: dict[str, dict[str, Any]] = {}
    scanned = 0
    usable = 0
    for rec in _scan_beneficiary_records(eff_cc, bundesland, fonds):
        scanned += 1
        place = _clean_location_value(rec["location"])
        if not place:
            continue
        usable += 1
        bucket = grouped.setdefault(place, {
            "value": 0.0, "count": 0, "fonds": set(), "bl": set(),
        })
        if rec["kosten"] is not None:
            bucket["value"] += float(rec["kosten"])
        bucket["count"] += 1
        if rec["fonds"]:
            bucket["fonds"].add(rec["fonds"])
        if rec["bundesland"]:
            bucket["bl"].add(rec["bundesland"])

    if not usable:
        return (
            f"Für {country_label} liegen in den geladenen Begünstigtenverzeichnissen "
            "keine auswertbaren Ortsangaben vor (die Listen führen überwiegend nur "
            "NUTS-Regionen oder Postleitzahlen)."
        )

    ranked = sorted(
        grouped.items(), key=lambda kv: (-float(kv[1]["value"] or 0.0), -kv[1]["count"])
    )[:10]
    lines = [f"Standorte mit dem höchsten Fördervolumen ({country_label}):"]
    for rank, (place, b) in enumerate(ranked, start=1):
        val = _format_eur_compact(b["value"])
        lines.append(f"{rank}. {place} — {val} · {_de_int(b['count'])} Vorhaben")
    lines.append("")
    lines.append(
        f"Datengrundlage: {country_label}, {_de_int(usable)} von {_de_int(scanned)} "
        "Vorhaben mit verwertbarem Ortsnamen."
    )
    return "\n".join(lines)


def _build_region_counts_answer(
    prompt: str,
    country_code: str | None,
    bundesland: str | None,
    fonds: str | None,
) -> str:
    """Anzahl geförderter Vorhaben pro Bundesland (zählen, nicht Volumen),
    aufgeschlüsselt nach Landesprogrammen (EFRE/ESF/JTF mit Länderbezug) vs.
    bundesweiten Programmen ohne Länderzuordnung (z. B. AMIF/ISF)."""
    eff_cc = _default_de_country(prompt, country_code)
    country_label = get_country_name(eff_cc) or eff_cc

    per_state: dict[str, dict[str, Any]] = {}
    federal = {"count": 0, "fonds": set()}
    total = 0
    for rec in _scan_beneficiary_records(eff_cc, bundesland, fonds):
        total += 1
        bl = (rec["bundesland"] or "").strip()
        if not bl or bl == "Unbekannt":
            federal["count"] += 1
            if rec["fonds"]:
                federal["fonds"].add(rec["fonds"])
            continue
        bucket = per_state.setdefault(bl, {"count": 0, "fonds": set()})
        bucket["count"] += 1
        if rec["fonds"]:
            bucket["fonds"].add(rec["fonds"])

    if not per_state and federal["count"] == 0:
        return (
            f"Für {country_label} liegen in den geladenen Begünstigtenverzeichnissen "
            "keine Vorhaben zum Zählen vor."
        )

    lines = [f"Anzahl geförderter Vorhaben pro Bundesland ({country_label}):"]
    if per_state:
        ranked = sorted(per_state.items(), key=lambda kv: (-kv[1]["count"], kv[0]))
        for rank, (bl, b) in enumerate(ranked, start=1):
            fonds_str = ", ".join(sorted(b["fonds"]))
            suffix = f" ({fonds_str})" if fonds_str else ""
            lines.append(f"{rank}. {bl} — {_de_int(b['count'])} Vorhaben{suffix}")
    if federal["count"] > 0:
        fed_fonds = ", ".join(sorted(federal["fonds"]))
        fed_hint = f" ({fed_fonds})" if fed_fonds else ""
        lines.append(
            f"Bundesweite Programme ohne Länderzuordnung: "
            f"{_de_int(federal['count'])} Vorhaben{fed_hint}."
        )
    lines.append(
        f"Datengrundlage: {country_label}, {_de_int(total)} Vorhaben gesamt, "
        f"{'1 Bundesland' if len(per_state) == 1 else f'{_de_int(len(per_state))} Bundesländer'}."
    )
    return "\n".join(lines)


def build_beneficiary_analysis_answer(
    prompt: str,
    limit: int = 5,
    country_code: str | None = None,
) -> str | None:
    sources = get_beneficiary_sources(country_code=country_code)
    if not sources:
        country_hint = ""
        if country_code:
            name = get_country_name(country_code) or country_code
            country_hint = f" für {name}"
        return f"Es sind derzeit keine Begünstigtenverzeichnisse{country_hint} geladen."

    matched_mode = _match_beneficiary_analysis_mode(prompt)
    if not matched_mode:
        return None

    bundesland, fonds = _extract_beneficiary_prompt_filters(prompt, sources)

    # Aggregat-Modi mit uneinheitlichen Quelldaten bekommen eine eigene, lesbare
    # Aufbereitung (vollständig, kompakte Beträge, bereinigt) statt der
    # generischen Top-5-Liste. Alle lesen bei jedem Aufruf live aus der DB und
    # spiegeln damit jeden neuen Begünstigten-Harvest.
    if matched_mode == "state_fund_totals":
        return _build_state_fund_totals_answer(prompt, country_code, bundesland, fonds)
    if matched_mode == "top_sectors":
        return _build_top_sectors_answer(prompt, country_code, bundesland, fonds)
    if matched_mode == "top_locations":
        return _build_top_locations_answer(prompt, country_code, bundesland, fonds)
    if matched_mode == "region_project_counts":
        return _build_region_counts_answer(prompt, country_code, bundesland, fonds)

    type_substrings, name_filter_label = _detect_beneficiary_name_filter(prompt)
    proper_nouns = _extract_prompt_proper_nouns(prompt)
    name_substrings: list[tuple[str, ...]] = []
    if type_substrings:
        name_substrings.append(type_substrings)
    for noun in proper_nouns:
        name_substrings.append((noun,))
    name_filter_parts: list[str] = []
    if name_filter_label:
        name_filter_parts.append(name_filter_label)
    if proper_nouns:
        name_filter_parts.append(" ".join(proper_nouns))
    name_filter_label_combined = " · ".join(name_filter_parts) if name_filter_parts else None

    # multi_state_beneficiaries: deutlich groesseres Limit, weil das eine
    # ueberschaubare Auswertung ist (typ. < 100 Einrichtungen) und der User
    # die volle Liste sehen will, nicht nur Top-X.
    if matched_mode == "multi_state_beneficiaries":
        effective_limit = max(limit, 200)
    else:
        effective_limit = max(3, min(limit, 8))

    analysis = analyze_beneficiary_records(
        mode=matched_mode,
        bundesland=bundesland,
        fonds=fonds,
        limit=effective_limit,
        country_code=country_code,
        name_substrings=name_substrings or None,
    )

    # 0-Treffer mit konkreten Eigennamen → LLM-Fallback erlauben
    # (z. B. „FH Gießen" → keine direkten Treffer, KI soll erkennen,
    # dass das die heutige Technische Hochschule Mittelhessen ist).
    if not analysis["items"] and (proper_nouns or name_filter_label):
        return None
    items = analysis["items"]
    summary = analysis["summary"]
    filters = analysis["filters"]

    if not items:
        missing_filters = [
            part
            for part in (
                f"Bundesland {filters['bundesland']}" if filters.get("bundesland") else None,
                f"Fonds {filters['fonds']}" if filters.get("fonds") else None,
                f"Begünstigte mit „{name_filter_label_combined}“" if name_filter_label_combined else None,
            )
            if part
        ]
        filter_note = f" für {' und '.join(missing_filters)}" if missing_filters else ""
        return (
            f"Für diese Fragestellung{filter_note} liegen in den aktuell geladenen "
            "Begünstigtenverzeichnissen keine passenden Auswertungsdaten vor."
        )

    intro_map = {
        "repeat_beneficiaries": "Die wichtigsten Begünstigten mit mehreren Vorhaben sind:",
        "multi_state_beneficiaries": (
            "Begünstigte, die in mehreren Bundesländer-Listen auftauchen "
            "(sortiert nach Anzahl der Bundesländer, dann nach Volumen):"
        ),
        "state_fund_totals": "Die höchsten Fördervolumina nach Bundesland und Fonds sind:",
        "top_locations": "Die Standorte mit dem höchsten Fördervolumen sind:",
        "top_beneficiaries": "Die größten Begünstigten sind:",
        "top_sectors": (
            "Die geförderten Wirtschaftszweige bzw. Interventionsbereiche "
            "(laut Spalten der Begünstigtenverzeichnisse) sind:"
        ),
    }
    intro = intro_map.get(matched_mode, analysis["title"] + ":")
    if matched_mode == "top_beneficiaries" and name_filter_label_combined:
        intro = f"Die größten Begünstigten mit „{name_filter_label_combined}“ sind:"
    lines = [intro]
    for item in items[:effective_limit]:
        val = item.get("value")
        value_display = (
            _format_eur_compact(val) if isinstance(val, (int, float)) and val
            else (item.get("value_label") or _format_eur(val))
        )
        line = f"{item['rank']}. {item['label']} — {value_display}"
        if item.get("project_count"):
            line += f" · {_de_int(item['project_count'])} Vorhaben"
        sub = item.get("sublabel") or ""
        if sub:
            parts = sub.split(" · ")
            # Führende "N Vorhaben"-Angabe entfernen — steht schon in der Zeile.
            if parts and re.search(r"\bVorhaben\b", parts[0]):
                parts = parts[1:]
            tail = " · ".join(p for p in parts if p)
            if tail:
                line += f" ({tail})"
        lines.append(line)

    coverage = (
        f"Datengrundlage: {_de_int(summary['sources_considered'])} Quelle(n), "
        f"{_de_int(summary['records_scanned'])} Datensätze, "
        f"{summary['total_volume_label']} Gesamtvolumen."
    )
    lines.append(coverage)
    if filters.get("bundesland") or filters.get("fonds") or name_filter_label_combined:
        active_filters = ", ".join(
            part
            for part in (
                f"Land={get_country_name(country_code) or country_code}" if country_code else None,
                f"Bundesland={filters['bundesland']}" if filters.get("bundesland") else None,
                f"Fonds={filters['fonds']}" if filters.get("fonds") else None,
                f"Filter={name_filter_label_combined}" if name_filter_label_combined else None,
            )
            if part
        )
        lines.append(f"Verwendete Filter: {active_filters}.")
    elif country_code:
        lines.append(f"Verwendeter Filter: Land={get_country_name(country_code) or country_code}.")
    return "\n".join(lines)


def _get_text_columns(source: str) -> list[str]:
    info = get_table_info(source)
    if not info.get("exists"):
        return []
    text_types = {"text", "character varying", "character"}
    return [c["name"] for c in info["columns"] if c["type"] in text_types]


def _pick_fallback_column(text_columns: list[str], used_columns: set[str]) -> str | None:
    for column in text_columns:
        if column not in used_columns:
            return column
    return None


def _get_column_names(source: str) -> list[str]:
    info = get_table_info(source)
    if not info.get("exists"):
        return []
    return [c["name"] for c in info["columns"]]


def _find_pattern_column(
    columns: list[str],
    patterns: list[str],
    used_columns: set[str] | None = None,
) -> str | None:
    blocked = used_columns or set()
    for pattern in patterns:
        for column in columns:
            if column in blocked:
                continue
            if re.search(pattern, column, re.IGNORECASE):
                return column
    return None


def _clean_display_text(value: Any) -> str:
    if value is None:
        return ""
    text_value = str(value).strip()
    if not text_value or text_value.lower() in {"nan", "none", "null"}:
        return ""
    return re.sub(r"\s+", " ", text_value)


def _format_reference_value(field_name: str, value: Any) -> str:
    if field_name == "amount":
        amount = _parse_numeric(value)
        if amount is not None:
            return _format_eur(amount)
    if isinstance(value, bool):
        return "Ja" if value else "Nein"
    return _clean_display_text(value)


def _join_unique_values(*values: Any, limit: int = 280) -> str:
    parts: list[str] = []
    seen: set[str] = set()
    for value in values:
        text_value = _clean_display_text(value)
        if not text_value:
            continue
        key = text_value.casefold()
        if key in seen:
            continue
        seen.add(key)
        parts.append(text_value)
    combined = " | ".join(parts)
    if len(combined) > limit:
        return combined[: limit - 1].rstrip() + "…"
    return combined


def _resolve_reference_registry_columns(source: str, registry_type: str | None) -> dict[str, str | None]:
    from services.geocoding_service import detect_columns

    columns = _get_column_names(source)
    text_columns = _get_text_columns(source)
    generic = detect_columns(source)
    profile_patterns = _REFERENCE_PROFILE_PATTERNS.get(registry_type or "", {})

    mapping = {
        "name": _find_pattern_column(columns, profile_patterns.get("name", [])) or generic.get("name"),
        "alias": _find_pattern_column(columns, profile_patterns.get("alias", [])),
        "projekt": _find_pattern_column(columns, profile_patterns.get("projekt", [])) or generic.get("projekt"),
        "beschreibung": _find_pattern_column(columns, profile_patterns.get("beschreibung", [])) or generic.get("beschreibung"),
        "aktenzeichen": _find_pattern_column(columns, profile_patterns.get("aktenzeichen", [])) or generic.get("aktenzeichen"),
        "standort": _find_pattern_column(columns, profile_patterns.get("standort", []))
        or generic.get("standort")
        or generic.get("ort")
        or generic.get("landkreis"),
        "country": _find_pattern_column(columns, profile_patterns.get("country", [])) or generic.get("country"),
        "status": _find_pattern_column(columns, profile_patterns.get("status", [])) or generic.get("status"),
        "authority": _find_pattern_column(columns, profile_patterns.get("authority", [])),
        "amount": _find_pattern_column(columns, profile_patterns.get("amount", [])),
        "programme": _find_pattern_column(columns, profile_patterns.get("programme", [])),
        "date": _find_pattern_column(columns, profile_patterns.get("date", [])),
    }

    used_for_fallback = {value for value in mapping.values() if value}
    if not mapping["name"]:
        mapping["name"] = _pick_fallback_column(text_columns, used_for_fallback)
        if mapping["name"]:
            used_for_fallback.add(mapping["name"])
    if not mapping["projekt"]:
        mapping["projekt"] = _pick_fallback_column(text_columns, used_for_fallback)
        if mapping["projekt"]:
            used_for_fallback.add(mapping["projekt"])
    if not mapping["beschreibung"]:
        mapping["beschreibung"] = _pick_fallback_column(text_columns, used_for_fallback)

    return mapping


def _build_reference_search_vectors(entry: dict[str, Any]) -> dict[str, str]:
    amount_value = _format_reference_value("amount", entry.get("amount"))
    return {
        "name": _join_unique_values(entry.get("name"), entry.get("alias")),
        "projekt": _join_unique_values(entry.get("projekt"), entry.get("programme")),
        "beschreibung": _join_unique_values(
            entry.get("beschreibung"),
            entry.get("authority"),
            amount_value,
            entry.get("date"),
        ),
        "aktenzeichen": _clean_display_text(entry.get("aktenzeichen")),
        "standort": _clean_display_text(entry.get("standort")),
        "country": _clean_display_text(entry.get("country")),
        "status": _join_unique_values(entry.get("status"), entry.get("date")),
    }


def _build_reference_project_name(entry: dict[str, Any], registry_type: str | None) -> str:
    fallback_order = {
        "sanctions": [entry.get("projekt"), entry.get("programme")],
        "tam": [entry.get("projekt"), entry.get("programme")],
        "state_aid": [entry.get("projekt"), entry.get("programme")],
        "cohesio": [entry.get("projekt"), entry.get("programme")],
    }
    return _join_unique_values(*(fallback_order.get(registry_type or "", [entry.get("projekt"), entry.get("programme")])), limit=180)


def _build_reference_description(entry: dict[str, Any], registry_type: str | None) -> str:
    ordered_fields = _REFERENCE_CONTEXT_ORDER.get(registry_type or "", _REFERENCE_CONTEXT_ORDER["other"])
    parts: list[str] = []
    for field_name in ordered_fields:
        formatted_value = _format_reference_value(field_name, entry.get(field_name))
        if not formatted_value:
            continue
        label = _REFERENCE_CONTEXT_LABELS.get(field_name)
        parts.append(f"{label}: {formatted_value}" if label else formatted_value)
    if not parts:
        return _clean_display_text(entry.get("beschreibung"))
    return _join_unique_values(*parts, limit=320)


def db_count_sources(
    country_code: str | None,
    bundesland: str | None,
    fonds: str | None,
    source: str | None,
) -> int:
    """Zaehlt distinkte source_keys in workshop_beneficiary_records, die zu
    den uebergebenen Filtern passen. Wird als ``sources_considered`` im
    Search-Response angezeigt.

    Wenn die Tabelle noch leer ist (frischer Container, noch kein Backfill),
    fallen wir transparent auf den Legacy-Counter aus
    ``get_beneficiary_sources()`` zurueck — sonst zeigt die UI dauerhaft 0.
    """
    where: list[str] = ["1=1"]
    params: dict[str, Any] = {}
    if country_code:
        where.append("country_code = :country_code")
        params["country_code"] = country_code.upper()
    if bundesland:
        where.append("bundesland = :bundesland")
        params["bundesland"] = bundesland
    if fonds:
        where.append("fonds = :fonds")
        params["fonds"] = fonds
    if source:
        where.append("source_key = :source")
        params["source"] = source
    sql = (
        "SELECT COUNT(DISTINCT source_key) FROM workshop_beneficiary_records "
        f"WHERE {' AND '.join(where)}"
    )
    try:
        with engine.connect() as conn:
            count = conn.execute(text(sql), params).scalar() or 0
        if count > 0:
            return int(count)
    except Exception as exc:  # noqa: BLE001
        log.debug("db_count_sources: zentrale Tabelle nicht abfragbar: %s", exc)
    # Fallback: Legacy-Quellenzaehlung auf workshop_df_metadata.
    legacy = [
        item for item in get_beneficiary_sources(country_code=country_code)
        if (not bundesland or (item.get("bundesland") or "") == bundesland)
        and (not fonds or (item.get("fonds") or "") == fonds)
        and (not source or item.get("source") == source)
    ]
    return len(legacy)


def search_beneficiary_records(
    query: str = "",
    scope: str = "all",
    bundesland: str | None = None,
    fonds: str | None = None,
    source: str | None = None,
    min_cost: float | None = None,
    limit: int = 60,
    company_limit: int = 14,
    country_code: str | None = None,
    min_score: float | None = None,
) -> dict:
    """Beguenstigtenverzeichnis-Suche mit rapidfuzz-Scoring (0..100).

    Phase 6a: Liest aus der zentralen Tabelle ``workshop_beneficiary_records``
    statt aus den 36+ per-Bundesland-Tabellen. SQL-Vorfilter via pg_trgm-GIN
    auf ``beneficiary_name_normalized`` und Token-ILIKE auf weitere Felder,
    danach rapidfuzz max(token_set_ratio, WRatio) auf das Kandidaten-Set.

    - ``min_score``: Fuzzy-Schwellwert pro Feld. Wenn None, adaptiv aus
      Query-Laenge (1 Token=80, 2=70, ≥3=60) — konsistent zur State-Aid-Suche.
    - Score-Skala: 0..100 (rapidfuzz max(token_set_ratio, WRatio)).
    - Treffer enthalten ``match_confidence`` (exact/high/medium/low). Das
      Legacy-Feld ``match_score_legacy`` wird mit 0/1 belegt und bleibt nur
      wegen Rueckwaertskompatibilitaet im Response — Werte sind nach dem
      Refactor nicht mehr auf der alten 0..140-Skala.
    """
    from services.company_aliases import expand_alias

    scope_fields = {
        "all": ["name", "projekt", "aktenzeichen", "standort", "beschreibung"],
        "company": ["name"],
        "project": ["projekt"],
        "aktenzeichen": ["aktenzeichen"],
        "location": ["standort"],
    }
    if scope not in scope_fields:
        raise ValueError(f"Unbekannter Scope '{scope}'.")

    # Alias-Expansion: 'KfW' -> 'Kreditanstalt für Wiederaufbau KfW'.
    effective_query, alias_label = expand_alias(query) if query else (query, None)
    normalized_query = _normalize_search_text(effective_query)

    # Adaptiver Schwellwert.
    auto_min_score = min_score is None
    effective_min_score = (
        float(min_score) if min_score is not None
        else _adaptive_min_score(effective_query or query or "")
    )

    # ── Innerer Scan: laeuft ggf. mehrfach mit gelockerten Filtern ──────────
    def _run_scan(
        active_bundesland: str | None,
        active_fonds: str | None,
        active_min_cost: float | None,
    ) -> tuple[list[dict[str, Any]], int, int]:
        # Filter-Bedingungen + Parameter zusammenbauen — alles auf der
        # zentralen Tabelle.
        where = ["1=1"]
        params: dict[str, Any] = {}
        if country_code:
            where.append("country_code = :country_code")
            params["country_code"] = country_code.upper()
        if active_bundesland:
            where.append("bundesland = :bundesland")
            params["bundesland"] = active_bundesland
        if active_fonds:
            where.append("fonds = :fonds")
            params["fonds"] = active_fonds
        if source:
            # Source-Argument bleibt erhalten — es entspricht dem Source-Key.
            where.append("source_key = :source")
            params["source"] = source
        if active_min_cost is not None:
            where.append("cost_total IS NOT NULL AND cost_total >= :min_cost")
            params["min_cost"] = active_min_cost

        # Token-Vorfilter: nur wenn die Query ueberhaupt Tokens hat. Sonst
        # darf der Aufrufer eine reine Filter-Suche fahren (Top-N pro
        # Bundesland o.ae.) — dann nehmen wir bis 5000 Records.
        # ESCAPE-Klausel weggelassen — PostgreSQL nutzt ohne ESCAPE-Klausel
        # standardmaessig den Backslash, was unsere ``escaped`` Strings
        # bereits beruecksichtigen. ESCAPE '\\' loest in psycopg2-Renderung
        # `'\\'` aus = zwei Zeichen → InvalidEscapeSequence.
        if normalized_query:
            tokens = [t for t in re.split(r"\s+", normalized_query) if len(t) >= 3]
            ilike_clauses: list[str] = []
            for i, tok in enumerate(tokens):
                key_raw = f"tok_{i}_raw"
                key_db = f"tok_{i}_db"
                escaped_raw = tok.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
                # Die Spalte beneficiary_name_normalized expandiert Umlaute
                # (ä→ae, ö→oe, ü→ue, ß→ss). Die Query-Normalisierung tut das
                # NICHT — also matcht ILIKE '%universität%' den DB-Wert
                # 'universitaet' nicht. Wir bauen daher zwei Varianten:
                # die Original-Form fuer project_name/location (nicht
                # umlaut-expandiert) und die ae/oe/ue/ss-Form fuer
                # beneficiary_name_normalized.
                tok_db = (
                    tok.replace("ä", "ae").replace("ö", "oe")
                    .replace("ü", "ue").replace("ß", "ss")
                )
                escaped_db = tok_db.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
                ilike_clauses.append(
                    f"(beneficiary_name_normalized ILIKE :{key_db} "
                    f"OR project_name ILIKE :{key_raw} "
                    f"OR project_aktenzeichen ILIKE :{key_raw} "
                    f"OR location ILIKE :{key_raw})"
                )
                params[key_raw] = f"%{escaped_raw}%"
                params[key_db] = f"%{escaped_db}%"
            if ilike_clauses:
                where.append("(" + " OR ".join(ilike_clauses) + ")")

        # NUTS-Code-Prefix — analog State-Aid: country_code kann genutzt
        # werden, um per Prefix-Match landesspezifisch zu filtern. Hier
        # dient country_code als Hauptfilter; Prefix-Match haengt mit
        # bundesland zusammen, dass NUTS-1 generiert wuerde — bewusst
        # nicht im Vorfilter, weil die Daten oft NULL haben.

        sql_filter = " AND ".join(where)
        # Bei reiner Filter-Suche koennen Tausende Records zurueckkommen —
        # wir limitieren das Kandidaten-Set hart auf 5000, damit rapidfuzz
        # nicht aus dem Ruder laeuft. Bei sehr grossen Bestaenden waere
        # eine `ORDER BY cost_total DESC NULLS LAST`-Heuristik denkbar,
        # ist aber bewusst weggelassen (Stable Behavior).
        sql_select = f"""
            SELECT id, source_key, bundesland, fonds, periode, country_code,
                   beneficiary_name, beneficiary_name_normalized,
                   project_name, project_aktenzeichen, project_description,
                   location, cost_total, source_filename, nuts_code
            FROM workshop_beneficiary_records
            WHERE {sql_filter}
            LIMIT 5000
        """

        with engine.connect() as conn:
            rows = conn.execute(text(sql_select), params).fetchall()

        results_local: list[dict[str, Any]] = []
        for row in rows:
            entry = dict(row._mapping)

            field_values = {
                "name": entry.get("beneficiary_name"),
                "projekt": entry.get("project_name"),
                "aktenzeichen": entry.get("project_aktenzeichen"),
                "standort": entry.get("location"),
                "beschreibung": entry.get("project_description"),
            }

            matched_fields: list[str] = []
            match_score: float = 0.0
            if normalized_query:
                for field_name in scope_fields[scope]:
                    field_value = field_values.get(field_name)
                    rf_score = _rapidfuzz_score(field_value, effective_query or query)
                    if rf_score >= effective_min_score:
                        matched_fields.append(field_name)
                        if rf_score > match_score:
                            match_score = rf_score
                if match_score < effective_min_score:
                    continue
            else:
                match_score = 1.0

            cost_value = float(entry["cost_total"]) if entry.get("cost_total") is not None else None
            country_name = get_country_name(entry.get("country_code"))

            results_local.append({
                "company_name": (entry.get("beneficiary_name") or "").strip() or "Unbekannt",
                "project_name": (entry.get("project_name") or "").strip(),
                "aktenzeichen": (entry.get("project_aktenzeichen") or "").strip(),
                "location": (entry.get("location") or "").strip(),
                "category": "",  # Phase 6a: kategorie nicht im zentralen Schema.
                "description": (entry.get("project_description") or "").strip(),
                "kosten": cost_value,
                "kosten_label": _format_eur(cost_value),
                "source": entry.get("source_key"),
                "bundesland": entry.get("bundesland"),
                "fonds": entry.get("fonds"),
                "periode": entry.get("periode"),
                "country_code": entry.get("country_code"),
                "country_name": country_name,
                # NUTS-Code mitfuehren — wird vom Audit-Report fuer Address-Match
                # genutzt (Polish-Runde 3, Mai 2026, Aufgabe 2).
                "nuts_code": (entry.get("nuts_code") or "").strip() or None,
                "matched_fields": matched_fields,
                "match_score": round(float(match_score), 1),
                "match_score_legacy": 0,
                "match_confidence": _match_confidence(match_score) if normalized_query else "exact",
            })

        # sources_considered = Anzahl distinkter source_keys, die nach Filter
        # uebrig bleiben (Vergleichbarkeit zur alten Logik).
        sources_in_scope = (
            db_count_sources(country_code, active_bundesland, active_fonds, source)
        )
        return results_local, len(rows), sources_in_scope

    # Empty-Result-Fallback (analog zur frueheren Logik).
    flat_results, scanned_records, sources_considered = _run_scan(bundesland, fonds, min_cost)
    relaxed: list[str] = []
    active_bundesland = bundesland
    active_fonds = fonds
    active_min_cost = min_cost

    if normalized_query and not flat_results:
        if active_bundesland:
            active_bundesland = None
            relaxed.append("bundesland")
            flat_results, scanned_records, sources_considered = _run_scan(
                active_bundesland, active_fonds, active_min_cost,
            )
        if not flat_results and active_min_cost is not None:
            active_min_cost = None
            relaxed.append("min_cost")
            flat_results, scanned_records, sources_considered = _run_scan(
                active_bundesland, active_fonds, active_min_cost,
            )
        if not flat_results and active_fonds:
            active_fonds = None
            relaxed.append("fonds")
            flat_results, scanned_records, sources_considered = _run_scan(
                active_bundesland, active_fonds, active_min_cost,
            )
        if not flat_results:
            relaxed = []

    def _record_sort_key(item: dict[str, Any]) -> tuple[Any, ...]:
        if normalized_query:
            return (
                -float(item.get("match_score") or 0.0),
                -float(item.get("kosten") or 0.0),
                item.get("company_name") or "",
                item.get("project_name") or "",
            )
        return (
            -float(item.get("kosten") or 0.0),
            item.get("company_name") or "",
            item.get("project_name") or "",
        )

    sorted_records = sorted(flat_results, key=_record_sort_key)

    companies: dict[str, dict[str, Any]] = {}
    for item in sorted_records:
        company_key = _normalize_search_text(item["company_name"]) or _normalize_search_text(item["project_name"])
        if not company_key:
            company_key = f"{item['source']}::{len(companies)}"
        company = companies.setdefault(company_key, {
            "company_name": item["company_name"],
            "total_kosten": 0.0,
            "project_count": 0,
            "match_score": 0.0,
            "match_score_legacy": 0,
            "sources": set(),
            "bundeslaender": set(),
            "fonds": set(),
            "standorte": set(),
            "aktenzeichen": set(),
            "matched_fields": set(),
            "projects": [],
        })

        if item.get("kosten") is not None:
            company["total_kosten"] += float(item["kosten"])
        company["project_count"] += 1
        company["match_score"] = max(
            company["match_score"], float(item.get("match_score") or 0.0),
        )
        company["match_score_legacy"] = max(
            company["match_score_legacy"], int(item.get("match_score_legacy") or 0),
        )
        if item.get("source"):
            company["sources"].add(item["source"])
        if item.get("bundesland"):
            company["bundeslaender"].add(item["bundesland"])
        if item.get("fonds"):
            company["fonds"].add(item["fonds"])
        if item.get("location"):
            company["standorte"].add(item["location"])
        if item.get("aktenzeichen"):
            company["aktenzeichen"].add(item["aktenzeichen"])
        for field_name in item.get("matched_fields") or []:
            company["matched_fields"].add(field_name)
        if len(company["projects"]) < 8:
            company["projects"].append({
                "project_name": item["project_name"],
                "aktenzeichen": item["aktenzeichen"],
                "location": item["location"],
                "category": item["category"],
                "kosten": item["kosten"],
                "kosten_label": item["kosten_label"],
                "source": item["source"],
                "bundesland": item["bundesland"],
                "fonds": item["fonds"],
                "periode": item["periode"],
                "country_code": item.get("country_code"),
                "country_name": item.get("country_name"),
                "matched_fields": item["matched_fields"],
                "match_score": item["match_score"],
                "match_score_legacy": item.get("match_score_legacy", 0),
                "match_confidence": item.get("match_confidence", "low"),
            })

    company_results = []
    for company in companies.values():
        company_results.append({
            "company_name": company["company_name"],
            "total_kosten": company["total_kosten"],
            "total_kosten_label": _format_eur(company["total_kosten"] or None),
            "project_count": company["project_count"],
            "match_score": round(float(company["match_score"]), 1),
            "match_score_legacy": company["match_score_legacy"],
            "match_confidence": _match_confidence(company["match_score"]) if normalized_query else "exact",
            "sources": sorted(company["sources"]),
            "bundeslaender": sorted(company["bundeslaender"]),
            "fonds": sorted(company["fonds"]),
            "standorte": sorted(company["standorte"])[:6],
            "aktenzeichen": sorted(company["aktenzeichen"])[:6],
            "matched_fields": sorted(company["matched_fields"]),
            "projects": company["projects"],
        })

    def _company_sort_key(item: dict[str, Any]) -> tuple[Any, ...]:
        if normalized_query:
            return (
                -float(item.get("match_score") or 0.0),
                -float(item.get("total_kosten") or 0.0),
                -int(item.get("project_count") or 0),
                item.get("company_name") or "",
            )
        return (
            -float(item.get("total_kosten") or 0.0),
            -int(item.get("project_count") or 0),
            item.get("company_name") or "",
        )

    company_results = sorted(company_results, key=_company_sort_key)
    limited_records = sorted_records[:max(1, min(limit, 200))]
    limited_companies = company_results[:max(1, min(company_limit, 200))]

    # Paket 2/3: Meta-Block — Alias-Expansion + gelockerte Filter sind fuer
    # den Aufrufer transparent (UI kann Hinweis-Banner anzeigen).
    meta: dict[str, Any] = {}
    if alias_label:
        meta["alias_used"] = alias_label
        meta["effective_query"] = effective_query
    if relaxed:
        meta["relaxed"] = relaxed
    if normalized_query:
        # Schwellwert + Skala transparent: 0..100 (rapidfuzz).
        meta["min_score"] = effective_min_score
        meta["score_scale"] = "0-100"
        if auto_min_score:
            meta["auto_min_score"] = True

    return {
        "query": query,
        "scope": scope,
        "summary": {
            "sources_considered": sources_considered,
            "records_scanned": scanned_records,
            "matches": len(sorted_records),
            "companies": len(company_results),
            "total_match_volume": sum(float(item["kosten"]) for item in sorted_records if item.get("kosten") is not None),
        },
        "meta": meta,
        "companies": limited_companies,
        "records": limited_records,
    }


def analyze_beneficiary_records(
    mode: str = "top_beneficiaries",
    bundesland: str | None = None,
    fonds: str | None = None,
    source: str | None = None,
    min_cost: float | None = None,
    limit: int = 10,
    country_code: str | None = None,
    name_substrings: tuple[str, ...] | list[tuple[str, ...]] | None = None,
) -> dict:
    from services.geocoding_service import detect_columns

    supported_modes = {
        "top_beneficiaries",
        "repeat_beneficiaries",
        "multi_state_beneficiaries",
        "state_fund_totals",
        "top_locations",
        "top_sectors",
        "region_project_counts",
        "kreis_project_counts",
    }
    if mode not in supported_modes:
        raise ValueError(f"Unbekannter Analysemodus '{mode}'.")

    # Bei multi_state_beneficiaries weicheres Cap (alle Einrichtungen mit
    # ≥2 Bundesländern sollen aufgelistet werden — typ. < 100 Einträge).
    # state_fund_totals kann durch alle BL/Fonds-Kombinationen pro Land 50+
    # Eintraege ausgeben — daher dort ebenfalls weicheres Cap.
    if mode == "multi_state_beneficiaries":
        max_limit = 500
    elif mode in {"state_fund_totals", "region_project_counts", "kreis_project_counts"}:
        max_limit = 100
    else:
        max_limit = 20
    limit = max(1, min(limit, max_limit))
    beneficiary_sources = [
        item for item in get_beneficiary_sources(country_code=country_code)
        if (not bundesland or (item.get("bundesland") or "") == bundesland)
        and (not fonds or (item.get("fonds") or "") == fonds)
        and (not source or item.get("source") == source)
    ]

    scanned_records = 0
    flat_results: list[dict[str, Any]] = []

    for source_info in beneficiary_sources:
        current_source = source_info["source"]
        table_name = _safe_table_name(current_source)
        columns = detect_columns(current_source)

        name_col = columns.get("name")
        project_col = columns.get("projekt")
        cost_col = columns.get("kosten")
        location_col = columns.get("standort") or columns.get("ort") or columns.get("landkreis")
        category_col = columns.get("sz")

        selected_cols: list[tuple[str, str]] = []
        for alias, col in (
            ("name", name_col),
            ("projekt", project_col),
            ("kosten", cost_col),
            ("standort", location_col),
            ("kategorie", category_col),
        ):
            if col:
                selected_cols.append((alias, col))

        if not selected_cols:
            continue

        sql = ", ".join(f"{_quote_ident(col)} AS {alias}" for alias, col in selected_cols)
        with engine.connect() as conn:
            rows = conn.execute(text(f'SELECT {sql} FROM "{table_name}"')).fetchall()

        scanned_records += len(rows)
        for row in rows:
            entry = dict(row._mapping)
            cost_value = _parse_numeric(entry.get("kosten"))
            if min_cost is not None and (cost_value is None or cost_value < min_cost):
                continue
            company_name_raw = str(entry.get("name") or "").strip()
            # Header-Artefakte aus XLSX-Erste-Zeilen rausfiltern
            if _is_header_artefact(company_name_raw):
                continue
            if name_substrings:
                lc = company_name_raw.casefold()
                # Eine Liste von Tupeln ⇒ AND zwischen Tupeln, OR innerhalb.
                # Ein einzelnes Tupel ⇒ nur OR (Backward-Compatibility).
                if name_substrings and isinstance(name_substrings[0], tuple):
                    if not all(
                        any(sub.casefold() in lc for sub in group)
                        for group in name_substrings
                    ):
                        continue
                elif not any(sub.casefold() in lc for sub in name_substrings):
                    continue

            flat_results.append({
                "company_name": company_name_raw or "Unbekannt",
                "project_name": str(entry.get("projekt") or "").strip() or "",
                "location": str(entry.get("standort") or "").strip() or "",
                "category": str(entry.get("kategorie") or "").strip() or "",
                "kosten": cost_value,
                "source": current_source,
                "bundesland": source_info.get("bundesland"),
                "fonds": source_info.get("fonds"),
                "periode": source_info.get("periode"),
                "country_code": source_info.get("country_code"),
                "country_name": source_info.get("country_name"),
            })

    total_volume = sum(float(item["kosten"]) for item in flat_results if item.get("kosten") is not None)
    items: list[dict[str, Any]] = []
    title = ""
    metric_label = "Fördervolumen"

    if mode in {"top_beneficiaries", "repeat_beneficiaries", "multi_state_beneficiaries"}:
        companies: dict[str, dict[str, Any]] = {}
        for item in flat_results:
            # Kanonischer Schluessel: Rechtsformsuffixe weg, Diakritika weg.
            # So fallen "Fraunhofer ... e.V.", "Fraunhofer ... e. V." und
            # "Fraunhofer ... eingetragener Verein" in denselben Bucket.
            company_key = (
                _canonical_company_key(item["company_name"])
                or _normalize_search_text(item["company_name"])
                or f"{item['source']}::{len(companies)}"
            )
            company = companies.setdefault(company_key, {
                "label": item["company_name"],
                "value": 0.0,
                "project_count": 0,
                "bundeslaender": set(),
                "fonds": set(),
                "sources": set(),
                "locations": set(),
            })
            if item.get("kosten") is not None:
                company["value"] += float(item["kosten"])
            company["project_count"] += 1
            if item.get("bundesland"):
                company["bundeslaender"].add(item["bundesland"])
            if item.get("fonds"):
                company["fonds"].add(item["fonds"])
            if item.get("source"):
                company["sources"].add(item["source"])
            if item.get("location"):
                company["locations"].add(item["location"])

        aggregated = []
        for company in companies.values():
            if mode == "repeat_beneficiaries" and company["project_count"] < 2:
                continue
            if mode == "multi_state_beneficiaries" and len(company["bundeslaender"]) < 2:
                continue
            aggregated.append({
                "label": company["label"],
                "sublabel": " · ".join(part for part in [
                    f"{_de_int(len(company['bundeslaender']))} Bundesländer" if mode == "multi_state_beneficiaries" else f"{_de_int(company['project_count'])} Vorhaben",
                    ", ".join(sorted(company["bundeslaender"])[:5]),
                    ", ".join(sorted(company["fonds"])[:2]),
                ] if part),
                "value": company["value"],
                "value_label": _format_eur(company["value"] or None),
                "project_count": company["project_count"],
                "source_count": len(company["sources"]),
                "sources": sorted(company["sources"]),
                "bundeslaender": sorted(company["bundeslaender"]),
                "bundeslaender_count": len(company["bundeslaender"]),
                "fonds_list": sorted(company["fonds"]),
                "locations": sorted(company["locations"])[:4],
            })

        if mode == "multi_state_beneficiaries":
            # Primaer nach Anzahl Bundeslaender sortieren, dann nach Volumen
            aggregated.sort(key=lambda item: (
                -int(item["bundeslaender_count"]),
                -float(item["value"] or 0.0),
                item["label"],
            ))
        else:
            aggregated.sort(key=lambda item: (
                -float(item["value"] or 0.0),
                -int(item["project_count"] or 0),
                item["label"],
            ))
        items = aggregated[:limit]
        if mode == "top_beneficiaries":
            title = "Größte Begünstigte"
        elif mode == "multi_state_beneficiaries":
            title = "Begünstigte in mehreren Bundesländern"
        else:
            title = "Begünstigte mit mehreren Vorhaben"

    elif mode == "state_fund_totals":
        grouped: dict[tuple[str, str], dict[str, Any]] = {}
        for item in flat_results:
            state = item.get("bundesland") or "Unbekannt"
            fund = item.get("fonds") or "Unbekannt"
            key = (state, fund)
            bucket = grouped.setdefault(key, {
                "label": f"{state} · {fund}",
                "value": 0.0,
                "project_count": 0,
                "bundesland": state,
                "fonds": fund,
                "sources": set(),
            })
            if item.get("kosten") is not None:
                bucket["value"] += float(item["kosten"])
            bucket["project_count"] += 1
            if item.get("source"):
                bucket["sources"].add(item["source"])

        items = sorted([
            {
                "label": bucket["label"],
                "sublabel": f"{_de_int(bucket['project_count'])} Vorhaben · {_de_int(len(bucket['sources']))} Quelle(n)",
                "value": bucket["value"],
                "value_label": _format_eur(bucket["value"] or None),
                "project_count": bucket["project_count"],
                "source_count": len(bucket["sources"]),
                "bundesland": bucket["bundesland"],
                "fonds": bucket["fonds"],
            }
            for bucket in grouped.values()
        ], key=lambda item: (-float(item["value"] or 0.0), item["label"]))[:limit]
        title = "Fördervolumen nach Bundesland und Fonds"

    elif mode == "top_sectors":
        # Wirtschaftszweig-/Interventionsbereich-Auswertung. Die `kategorie`-
        # Spalte stammt aus geocoding_service.detect_columns(...).get("sz")
        # und faengt Wirtschaftstaetigkeit, Art der Intervention,
        # Interventionsbereich/-kategorie sowie Spezifisches Ziel ab.
        grouped_sectors: dict[str, dict[str, Any]] = {}
        records_with_category = 0
        for item in flat_results:
            category = (item.get("category") or "").strip()
            if not category:
                continue
            records_with_category += 1
            # Fuer lange Klartext-Kategorien auf eine handhabbare Laenge kuerzen
            label = category if len(category) <= 110 else category[:107].rstrip() + "…"
            bucket = grouped_sectors.setdefault(label, {
                "label": label,
                "value": 0.0,
                "project_count": 0,
                "bundeslaender": set(),
                "fonds": set(),
            })
            if item.get("kosten") is not None:
                bucket["value"] += float(item["kosten"])
            bucket["project_count"] += 1
            if item.get("bundesland"):
                bucket["bundeslaender"].add(item["bundesland"])
            if item.get("fonds"):
                bucket["fonds"].add(item["fonds"])

        items = sorted([
            {
                "label": bucket["label"],
                "sublabel": " · ".join(part for part in [
                    f"{_de_int(bucket['project_count'])} Vorhaben",
                    ", ".join(sorted(bucket["bundeslaender"])[:3]),
                    ", ".join(sorted(bucket["fonds"])[:2]),
                ] if part),
                "value": bucket["value"],
                "value_label": _format_eur(bucket["value"] or None),
                "project_count": bucket["project_count"],
                "bundeslaender": sorted(bucket["bundeslaender"]),
                "fonds_list": sorted(bucket["fonds"]),
            }
            for bucket in grouped_sectors.values()
        ], key=lambda item: (-float(item["value"] or 0.0), -int(item["project_count"]), item["label"]))[:limit]
        title = (
            f"Wirtschaftszweige / Interventionsbereiche "
            f"({_de_int(records_with_category)} von {_de_int(scanned_records)} Vorhaben mit Angabe)"
        )

    elif mode == "top_locations":
        grouped_locations: dict[str, dict[str, Any]] = {}
        for item in flat_results:
            location = item.get("location") or ""
            if not location:
                continue
            bucket = grouped_locations.setdefault(location, {
                "label": location,
                "value": 0.0,
                "project_count": 0,
                "bundeslaender": set(),
                "fonds": set(),
            })
            if item.get("kosten") is not None:
                bucket["value"] += float(item["kosten"])
            bucket["project_count"] += 1
            if item.get("bundesland"):
                bucket["bundeslaender"].add(item["bundesland"])
            if item.get("fonds"):
                bucket["fonds"].add(item["fonds"])

        items = sorted([
            {
                "label": bucket["label"],
                "sublabel": " · ".join(part for part in [
                    f"{_de_int(bucket['project_count'])} Vorhaben",
                    ", ".join(sorted(bucket["bundeslaender"])[:2]),
                    ", ".join(sorted(bucket["fonds"])[:2]),
                ] if part),
                "value": bucket["value"],
                "value_label": _format_eur(bucket["value"] or None),
                "project_count": bucket["project_count"],
                "bundeslaender": sorted(bucket["bundeslaender"]),
                "fonds_list": sorted(bucket["fonds"]),
            }
            for bucket in grouped_locations.values()
        ], key=lambda item: (-float(item["value"] or 0.0), item["label"]))[:limit]
        title = "Standorte mit dem höchsten Fördervolumen"

    elif mode == "region_project_counts":
        # Vorhaben je Bundesland mit Quellen-/Fonds-Aufschluesselung.
        # Aggregation aus dem bereits gefilterten flat_results-Set; pro
        # Bundesland-Bucket eine zweite Sub-Aggregation nach
        # (source, fonds), damit der Pruefer sieht, aus welchem Topf
        # die Vorhaben stammen.
        grouped_regions: dict[str, dict[str, Any]] = {}
        for item in flat_results:
            state = item.get("bundesland") or "Unbekannt"
            bucket = grouped_regions.setdefault(state, {
                "label": state,
                "value": 0.0,
                "project_count": 0,
                "bundesland": state,
                "sources": set(),
                "fonds": set(),
                "sub_buckets": {},
            })
            if item.get("kosten") is not None:
                bucket["value"] += float(item["kosten"])
            bucket["project_count"] += 1
            src_key = item.get("source") or "unknown"
            fonds_label = item.get("fonds") or "Unbekannt"
            if item.get("source"):
                bucket["sources"].add(item["source"])
            if item.get("fonds"):
                bucket["fonds"].add(item["fonds"])
            sub_key = (src_key, fonds_label)
            sub = bucket["sub_buckets"].setdefault(sub_key, {
                "source": src_key,
                "fonds": fonds_label,
                "count": 0,
                "value": 0.0,
            })
            sub["count"] += 1
            if item.get("kosten") is not None:
                sub["value"] += float(item["kosten"])

        items_local: list[dict[str, Any]] = []
        for bucket in grouped_regions.values():
            sources_breakdown = sorted(
                (
                    {
                        "source": sub["source"],
                        "fonds": sub["fonds"],
                        "count": sub["count"],
                        "value": sub["value"],
                        "value_label": _format_eur(sub["value"] or None),
                    }
                    for sub in bucket["sub_buckets"].values()
                ),
                key=lambda s: (-int(s["count"]), s["source"]),
            )
            items_local.append({
                "label": bucket["label"],
                "sublabel": (
                    f"{_de_int(bucket['project_count'])} Vorhaben · "
                    f"{_de_int(len(bucket['sources']))} Quelle(n)"
                ),
                "value": bucket["project_count"],
                "value_label": f"{_de_int(bucket['project_count'])} Vorhaben",
                "project_count": bucket["project_count"],
                "bundesland": bucket["bundesland"],
                "fonds": None,
                "source_count": len(bucket["sources"]),
                "fonds_list": sorted(bucket["fonds"]),
                "total_volume": bucket["value"],
                "total_volume_label": _format_eur(bucket["value"] or None),
                "sources_breakdown": sources_breakdown,
            })
        items = sorted(
            items_local,
            key=lambda it: (-int(it["project_count"] or 0), it["label"]),
        )[:limit]
        title = "Vorhaben je Bundesland (mit Quellen-Aufschlüsselung)"
        metric_label = "Vorhaben"

    elif mode == "kreis_project_counts":
        # Vorhaben je Kreis (NUTS-3). Liest direkt aus
        # workshop_beneficiary_records, weil nur dort der nuts_code-
        # Spalte verlaesslich gefuellt ist; die XLSX-Quellen
        # (get_beneficiary_sources) haben kein NUTS-Feld im
        # detect_columns-Schema.
        from services.geocoding_service import lookup_nuts_code  # lazy import

        kreis_where: list[str] = ["nuts_code IS NOT NULL", "nuts_code <> ''"]
        kreis_params: dict[str, Any] = {}
        if country_code:
            kreis_where.append("country_code = :country_code")
            kreis_params["country_code"] = country_code.upper()
        if bundesland:
            kreis_where.append("bundesland = :bundesland")
            kreis_params["bundesland"] = bundesland
        if fonds:
            kreis_where.append("fonds = :fonds")
            kreis_params["fonds"] = fonds
        if source:
            kreis_where.append("source_key = :source")
            kreis_params["source"] = source
        if min_cost is not None:
            kreis_where.append("cost_total IS NOT NULL AND cost_total >= :min_cost")
            kreis_params["min_cost"] = float(min_cost)

        kreis_sql = (
            "SELECT nuts_code, source_key, fonds, bundesland, "
            "       COUNT(*) AS cnt, "
            "       COALESCE(SUM(cost_total), 0) AS sum_cost "
            "FROM workshop_beneficiary_records "
            f"WHERE {' AND '.join(kreis_where)} "
            "GROUP BY nuts_code, source_key, fonds, bundesland"
        )
        kreis_grouped: dict[str, dict[str, Any]] = {}
        kreis_scanned = 0
        try:
            with engine.connect() as conn:
                rows = conn.execute(text(kreis_sql), kreis_params).fetchall()
        except Exception as exc:  # noqa: BLE001
            log.warning("kreis_project_counts: DB-Query fehlgeschlagen: %s", exc)
            rows = []

        for row in rows:
            r = dict(row._mapping)
            nuts = (r.get("nuts_code") or "").strip().upper()
            if not nuts:
                continue
            cnt = int(r.get("cnt") or 0)
            kreis_scanned += cnt
            sum_cost = float(r.get("sum_cost") or 0.0)
            bucket = kreis_grouped.setdefault(nuts, {
                "nuts_code": nuts,
                "label": nuts,
                "bundesland": r.get("bundesland") or "",
                "value": 0.0,
                "project_count": 0,
                "sources": set(),
                "fonds": set(),
                "sub_buckets": {},
            })
            bucket["project_count"] += cnt
            bucket["value"] += sum_cost
            src_key = r.get("source_key") or "unknown"
            fonds_label = r.get("fonds") or "Unbekannt"
            if r.get("source_key"):
                bucket["sources"].add(r["source_key"])
            if r.get("fonds"):
                bucket["fonds"].add(r["fonds"])
            if not bucket["bundesland"] and r.get("bundesland"):
                bucket["bundesland"] = r["bundesland"]
            sub_key = (src_key, fonds_label)
            sub = bucket["sub_buckets"].setdefault(sub_key, {
                "source": src_key,
                "fonds": fonds_label,
                "count": 0,
                "value": 0.0,
            })
            sub["count"] += cnt
            sub["value"] += sum_cost

        items_local: list[dict[str, Any]] = []
        for bucket in kreis_grouped.values():
            nuts = bucket["nuts_code"]
            info = lookup_nuts_code(nuts) or {}
            kreis_name = info.get("ort") or nuts
            bl = info.get("bundesland") or bucket.get("bundesland") or ""
            display_label = (
                f"{kreis_name} ({nuts})" if kreis_name and kreis_name != nuts else nuts
            )
            sources_breakdown = sorted(
                (
                    {
                        "source": sub["source"],
                        "fonds": sub["fonds"],
                        "count": sub["count"],
                        "value": sub["value"],
                        "value_label": _format_eur(sub["value"] or None),
                    }
                    for sub in bucket["sub_buckets"].values()
                ),
                key=lambda s: (-int(s["count"]), s["source"]),
            )
            items_local.append({
                "label": display_label,
                "sublabel": (
                    f"{_de_int(bucket['project_count'])} Vorhaben · "
                    f"{_de_int(len(bucket['sources']))} Quelle(n)"
                    + (f" · {bl}" if bl else "")
                ),
                "value": bucket["project_count"],
                "value_label": f"{_de_int(bucket['project_count'])} Vorhaben",
                "project_count": bucket["project_count"],
                "nuts_code": nuts,
                "bundesland": bl,
                "fonds": None,
                "source_count": len(bucket["sources"]),
                "fonds_list": sorted(bucket["fonds"]),
                "total_volume": bucket["value"],
                "total_volume_label": _format_eur(bucket["value"] or None),
                "sources_breakdown": sources_breakdown,
            })
        items = sorted(
            items_local,
            key=lambda it: (-int(it["project_count"] or 0), it["label"]),
        )[:limit]
        # scanned_records bleibt bei 0 fuer den XLSX-Pfad — wir setzen
        # einen sinnvollen Ersatzwert aus dem DB-Aggregat, damit das
        # Coverage-Summary nicht 0 anzeigt.
        if kreis_scanned and not scanned_records:
            scanned_records = kreis_scanned
        title = "Vorhaben je Kreis (NUTS-3, mit Quellen-Aufschlüsselung)"
        metric_label = "Vorhaben"

    for index, item in enumerate(items, start=1):
        item["rank"] = index

    return {
        "mode": mode,
        "title": title,
        "metric_label": metric_label,
        "summary": {
            "sources_considered": len(beneficiary_sources),
            "records_scanned": scanned_records,
            "items": len(items),
            "total_volume": total_volume,
            "total_volume_label": _format_eur(total_volume or None),
        },
        "filters": {
            "bundesland": bundesland,
            "fonds": fonds,
            "source": source,
            "min_cost": min_cost,
            "country_code": country_code,
        },
        "items": items,
    }


def search_reference_registry_records(
    query: str,
    registry_type: str | None = None,
    source: str | None = None,
    limit: int = 30,
) -> dict:
    normalized_query = _normalize_search_text(query)
    registry_sources = [
        item for item in list_reference_registry_sources()
        if (not registry_type or item.get("registry_type") == registry_type)
        and (not source or item.get("source") == source)
    ]

    if not normalized_query:
        return {
            "query": query,
            "summary": {
                "sources_considered": len(registry_sources),
                "matches": 0,
            },
            "hits": [],
        }

    hits: list[dict[str, Any]] = []

    for source_info in registry_sources:
        current_source = source_info["source"]
        table_name = _safe_table_name(current_source)
        current_registry_type = source_info.get("registry_type") or "other"
        columns = _resolve_reference_registry_columns(current_source, current_registry_type)

        selected_cols: list[tuple[str, str]] = []
        for alias, col in (
            ("name", columns.get("name")),
            ("alias", columns.get("alias")),
            ("projekt", columns.get("projekt")),
            ("beschreibung", columns.get("beschreibung")),
            ("aktenzeichen", columns.get("aktenzeichen")),
            ("standort", columns.get("standort")),
            ("country", columns.get("country")),
            ("status", columns.get("status")),
            ("authority", columns.get("authority")),
            ("amount", columns.get("amount")),
            ("programme", columns.get("programme")),
            ("date", columns.get("date")),
        ):
            if col:
                selected_cols.append((alias, col))

        if not selected_cols:
            continue

        sql = ", ".join(f"{_quote_ident(col)} AS {alias}" for alias, col in selected_cols)
        with engine.connect() as conn:
            rows = conn.execute(text(f'SELECT {sql} FROM "{table_name}" LIMIT 5000')).fetchall()

        for row in rows:
            entry = dict(row._mapping)
            field_values = _build_reference_search_vectors(entry)

            match_score = 0
            matched_fields: list[str] = []
            for field_name in ("name", "projekt", "beschreibung", "aktenzeichen", "standort", "country", "status"):
                field_score = _score_search_value(field_values.get(field_name), normalized_query)
                if field_score > 0:
                    matched_fields.append(field_name)
                    match_score = max(match_score, field_score)

            if match_score == 0:
                continue

            hits.append({
                "company_name": _join_unique_values(entry.get("name"), entry.get("alias"), limit=160) or _build_reference_project_name(entry, current_registry_type) or "Unbekannt",
                "project_name": _build_reference_project_name(entry, current_registry_type),
                "description": _build_reference_description(entry, current_registry_type),
                "aktenzeichen": _clean_display_text(entry.get("aktenzeichen")),
                "location": _clean_display_text(entry.get("standort")),
                "country": _clean_display_text(entry.get("country")),
                "status": _clean_display_text(entry.get("status")),
                "source": current_source,
                "registry_type": current_registry_type,
                "filename": source_info.get("filename"),
                "matched_fields": matched_fields,
                "match_score": match_score,
            })

    total_matches = len(hits)
    hits = sorted(
        hits,
        key=lambda item: (
            -int(item.get("match_score") or 0),
            item.get("registry_type") or "",
            item.get("company_name") or "",
            item.get("project_name") or "",
        ),
    )[:max(1, min(limit, 100))]

    return {
        "query": query,
        "summary": {
            "sources_considered": len(registry_sources),
            "matches": total_matches,
        },
        "hits": hits,
    }


def list_dataframe_tables() -> list[dict]:
    """Listet alle workshop_df_* Tabellen."""
    _ensure_metadata_table()
    with engine.connect() as conn:
        result = conn.execute(text("""
            SELECT
                t.table_name,
                m.dataset_group,
                m.registry_type
            FROM information_schema.tables t
            LEFT JOIN workshop_df_metadata m ON m.table_name = t.table_name
            WHERE t.table_schema = 'public' AND t.table_name LIKE 'workshop_df_%'
            ORDER BY t.table_name
        """))
        tables = []
        for name, dataset_group, registry_type in result:
            source = name.replace("workshop_df_", "")
            count = conn.execute(text(f'SELECT COUNT(*) FROM "{name}"')).scalar()
            tables.append({
                "table_name": name,
                "source": source,
                "rows": count,
                "dataset_group": dataset_group,
                "registry_type": registry_type,
            })
    return tables


def delete_dataframe_table(source: str) -> bool:
    """Loescht eine DataFrame-Tabelle und ihre Metadaten."""
    table_name = _safe_table_name(source)
    with engine.connect() as conn:
        conn.execute(text(f'DROP TABLE IF EXISTS "{table_name}"'))
        conn.execute(text("DELETE FROM workshop_df_metadata WHERE table_name = :t OR source = :s"),
                     {"t": table_name, "s": source})
        conn.commit()
    log.info("DataFrame-Tabelle %s + Metadaten geloescht.", table_name)
    return True
