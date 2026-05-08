"""
flowworkshop · services/excel_export.py

Zentraler XLSX-Export-Helper. Erzeugt aus einer Liste von Dict-Datensaetzen
eine fertig formatierte XLSX-Datei (Bytes), die in Excel/LibreOffice mit
Filtern, Header-Formatierung und einem Pflicht-/Datenstand-Hinweis-Sheet
geoeffnet werden kann.

Verwendet von:
- routers/state_aid.py        (Award-Liste, Stats, Audit-Trail)
- routers/sanctions.py        (Treffer-Liste)
- routers/beneficiaries.py    (Beneficiary-Search)

Backend-Konvention: Pflichthinweise werden in einem zusaetzlichen Sheet
"Hinweise" abgelegt — Excel-Filter sollen das Datensheet nicht blockieren,
deshalb steht der Hinweistext nicht im selben Sheet wie die Daten.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet


# Header-Style (konsistent ueber alle Exporte)
_HEADER_FILL = PatternFill("solid", fgColor="0E7C5C")  # Workshop-Emerald
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(vertical="center", wrap_text=True)

# Alternierende Zeilenfarbe — mildes Grau
_ROW_FILL = PatternFill("solid", fgColor="F1F5F9")

_NOTES_TITLE_FONT = Font(bold=True, size=12, color="0E1F18")
_NOTES_BODY_FONT = Font(size=10, color="334155")


def _coerce_cell(value: Any) -> Any:
    """Wandelt Werte in einen Excel-kompatiblen Typ um.

    - datetime/date passieren durch (openpyxl rendert sie nativ)
    - dict/list werden joined (z.B. Aliase, NUTS-Listen)
    - None wird zu "" damit das Sheet sauber aussieht
    """
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool, datetime)):
        return value
    if isinstance(value, (list, tuple, set)):
        return " | ".join(str(v) for v in value if v is not None)
    if isinstance(value, dict):
        return " | ".join(f"{k}={v}" for k, v in value.items())
    # Decimals / Date / Sonstiges → str
    return str(value)


def _autosize_columns(ws: Worksheet, headers: Sequence[str], rows: Sequence[Sequence[Any]],
                      max_width: int = 60) -> None:
    """Setzt eine grobe Auto-Width pro Spalte.

    openpyxl kennt keine echte Auto-Width — wir nehmen das Maximum aus
    Header-Laenge und den ersten 200 Daten-Zeilen.
    """
    sample = rows[:200]
    for idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in sample:
            if idx - 1 < len(row):
                v = row[idx - 1]
                if v is None:
                    continue
                length = len(str(v))
                if length > max_len:
                    max_len = length
        # 1.2 Faktor weil openpyxl-Width != Pixel
        width = min(max_width, max(10, int(max_len * 1.15) + 2))
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_data_sheet(
    ws: Worksheet,
    *,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    table_name: str | None = None,
) -> None:
    """Schreibt Daten + Header + Auto-Width + AutoFilter + alternierende Farben.

    Wenn ``table_name`` gesetzt ist, wird ein Excel-Table-Objekt angelegt —
    das aktiviert in Excel das Filter-Dropdown und die Sortier-Knoepfe.
    """
    # Header-Zeile
    ws.append(list(headers))
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    ws.row_dimensions[1].height = 22

    # Daten-Zeilen
    for row_idx, row in enumerate(rows, start=2):
        ws.append([_coerce_cell(v) for v in row])
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = _ROW_FILL

    # Spalten-Breite
    _autosize_columns(ws, headers, rows)

    # AutoFilter / Excel-Table
    last_col = get_column_letter(len(headers))
    last_row = max(2, len(rows) + 1)
    ref = f"A1:{last_col}{last_row}"

    if table_name and len(rows) > 0:
        # Excel-Tables kosten nichts und liefern Filter + Banding gratis
        try:
            tbl = Table(displayName=_safe_table_name(table_name), ref=ref)
            tbl.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showRowStripes=True,
            )
            ws.add_table(tbl)
        except ValueError:
            # ungueltiger Name → fallback nur AutoFilter
            ws.auto_filter.ref = ref
    else:
        ws.auto_filter.ref = ref

    # Erste Zeile fixieren
    ws.freeze_panes = "A2"


def _safe_table_name(name: str) -> str:
    """Excel-Table-Namen muessen mit Buchstaben/Underscore beginnen, keine Spaces."""
    safe = "".join(c if c.isalnum() else "_" for c in name)
    if not safe or not (safe[0].isalpha() or safe[0] == "_"):
        safe = "Tab_" + safe
    return safe[:240] or "Tab"


def write_notes_sheet(
    ws: Worksheet,
    *,
    title: str,
    pflichthinweis: str,
    metadata: dict[str, Any] | None = None,
    additional_lines: Sequence[str] | None = None,
) -> None:
    """Schreibt das ``Hinweise``-Sheet mit Pflichthinweis + Metadaten.

    Layout (Spalten A/B):
      A1: Titel (fett, groesser)
      A3..: Schluessel  | Wert
    """
    ws["A1"] = title
    ws["A1"].font = _NOTES_TITLE_FONT
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 90
    ws["A2"] = ""

    row = 3
    ws.cell(row=row, column=1, value="Pflichthinweis").font = Font(bold=True, size=10)
    cell = ws.cell(row=row, column=2, value=pflichthinweis)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.font = _NOTES_BODY_FONT
    ws.row_dimensions[row].height = 90
    row += 2

    ws.cell(row=row, column=1, value="Erzeugt").font = Font(bold=True, size=10)
    ws.cell(row=row, column=2, value=datetime.utcnow().isoformat() + "Z").font = _NOTES_BODY_FONT
    row += 1

    if metadata:
        for k, v in metadata.items():
            ws.cell(row=row, column=1, value=str(k)).font = Font(bold=True, size=10)
            cell = ws.cell(row=row, column=2, value=_coerce_cell(v))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = _NOTES_BODY_FONT
            row += 1

    if additional_lines:
        row += 1
        for line in additional_lines:
            cell = ws.cell(row=row, column=1, value=line)
            cell.font = _NOTES_BODY_FONT
            cell.alignment = Alignment(wrap_text=True)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            row += 1


def make_xlsx(
    rows: Iterable[dict[str, Any]],
    *,
    sheet_name: str,
    headers: Sequence[str],
    table_name: str | None = None,
    pflichthinweis: str | None = None,
    metadata: dict[str, Any] | None = None,
    notes_title: str = "Hinweise zum Export",
    additional_notes: Sequence[str] | None = None,
) -> bytes:
    """Erzeugt eine vollstaendige XLSX-Datei (Bytes) aus einer Datensatz-Liste.

    Kapselt das Pattern aus `services/state_aid_audit_pdf.py`:
    - Sheet 1: Daten (Header fett, AutoFilter, Auto-Width, alternierende Farben)
    - Sheet "Hinweise": Pflichthinweis, Metadaten, Datenstand pro Quelle.

    Parameters
    ----------
    rows
        Liste von Dicts. Jeder Dict-Schluessel sollte in ``headers`` enthalten
        sein — fehlende Schluessel werden als leer behandelt.
    sheet_name
        Name des Daten-Sheets (Excel max. 31 Zeichen).
    headers
        Reihenfolge der Spalten im Daten-Sheet. Bestimmt auch die
        Reihenfolge in der Header-Zeile.
    table_name
        Optionaler Name fuer das Excel-Table-Objekt. Sorgt fuer Filter-
        Dropdown und Banding. Whitespace + Sonderzeichen werden bereinigt.
    pflichthinweis
        Datenherkunfts-Hinweis (Plan §13). Wenn None, wird das Hinweise-Sheet
        weggelassen.
    metadata
        Optionales Dict (z.B. Filter, Datenstand pro Quelle). Wird im
        Hinweise-Sheet als Schluessel-Wert-Tabelle ausgegeben.
    additional_notes
        Optionale freitextliche Zeilen unter den Metadaten (z.B. Quellen-
        Liste pro Quelle mit Datum).
    """
    rows_list = list(rows)
    serialized_rows: list[list[Any]] = []
    for record in rows_list:
        serialized_rows.append([record.get(h) for h in headers])

    wb = Workbook()
    ws_data = wb.active
    if ws_data is None:
        ws_data = wb.create_sheet()
    ws_data.title = (sheet_name or "Daten")[:31]
    write_data_sheet(
        ws_data,
        headers=headers,
        rows=serialized_rows,
        table_name=table_name,
    )

    if pflichthinweis:
        ws_notes = wb.create_sheet("Hinweise")
        write_notes_sheet(
            ws_notes,
            title=notes_title,
            pflichthinweis=pflichthinweis,
            metadata=metadata,
            additional_lines=additional_notes,
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_xlsx_multi_sheet(
    sheets: Sequence[dict[str, Any]],
    *,
    pflichthinweis: str | None = None,
    metadata: dict[str, Any] | None = None,
    notes_title: str = "Hinweise zum Export",
    additional_notes: Sequence[str] | None = None,
) -> bytes:
    """Erzeugt eine XLSX-Datei mit mehreren Sheets (z.B. Statistiken).

    ``sheets`` ist eine Liste von Dicts mit den Schluesseln:
      - name        Sheet-Name
      - headers     Spalten-Reihenfolge
      - rows        Liste von Dicts (rohe Datensaetze)
      - table_name  optional, fuer das Excel-Table-Objekt
    """
    wb = Workbook()
    # default-Sheet nicht direkt verwenden — sonst landet der erste Tab
    # immer mit Default-Namen "Sheet" am Anfang
    default = wb.active
    if default is not None:
        wb.remove(default)

    for sheet_def in sheets:
        name = (sheet_def.get("name") or "Daten")[:31]
        headers = list(sheet_def.get("headers") or [])
        rows = sheet_def.get("rows") or []
        ws = wb.create_sheet(name)
        serialized = [[row.get(h) for h in headers] for row in rows]
        write_data_sheet(
            ws,
            headers=headers,
            rows=serialized,
            table_name=sheet_def.get("table_name"),
        )

    if pflichthinweis:
        ws_notes = wb.create_sheet("Hinweise")
        write_notes_sheet(
            ws_notes,
            title=notes_title,
            pflichthinweis=pflichthinweis,
            metadata=metadata,
            additional_lines=additional_notes,
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── HTTP-Helfer ──────────────────────────────────────────────────────────────


def xlsx_response_headers(filename: str) -> dict[str, str]:
    """Standard-Header fuer XLSX-Downloads (Content-Disposition + nosniff)."""
    safe = "".join(c if (c.isalnum() or c in "._-") else "_" for c in filename)
    if not safe.lower().endswith(".xlsx"):
        safe += ".xlsx"
    return {
        "Content-Disposition": f'attachment; filename="{safe}"',
        "X-Content-Type-Options": "nosniff",
    }


XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
