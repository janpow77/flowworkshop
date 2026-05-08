#!/usr/bin/env python3
"""
flowworkshop · scripts/harvest_beneficiary_url.py

Wiederverwendbarer Beneficiary-Harvester: laedt eine xlsx-Liste von einer
URL und uebergibt sie an den /api/beneficiaries/upload-Endpoint mit
mode=force, sodass eine bestehende Quelle ueberschrieben wird.

Ist gedacht fuer offizielle Listen-Updates der Bundeslaender, die ihre
Beguenstigtenverzeichnisse als XLSX am gleichen URL veroeffentlichen
(Beispiele Hessen ESF, Brandenburg EFRE/JTF, …). Der Worker-Token wird
aus der Backend-Config gelesen.

Beispiele:

    # Hessen ESF (komplette Datei, eine Source):
    python scripts/harvest_beneficiary_url.py \\
      --url "https://www.esf-hessen.de/resource/blob/esf-hessen/603958/.../liste-der-vorhaben-2021-2027-esf--data.xlsx"

    # Brandenburg (kombinierte EFRE+JTF-Datei, splitten nach Fonds-Spalte):
    python scripts/harvest_beneficiary_url.py \\
      --url "https://efre.brandenburg.de/sixcms/media.php/9/2025_12_31%20Liste%20der%20Vorhaben.xlsx" \\
      --split-fonds-column 11

    # Trockenlauf — nur herunterladen + Spalten-Sample, keinen Upload:
    python scripts/harvest_beneficiary_url.py --url "..." --dry
"""
from __future__ import annotations

import argparse
import io
import logging
import os
import sys
import tempfile
from pathlib import Path
from typing import Optional

import httpx

logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
log = logging.getLogger("harvest_beneficiary_url")


def download(url: str) -> bytes:
    log.info("Lade XLSX von %s", url)
    with httpx.Client(timeout=60.0, follow_redirects=True) as client:
        resp = client.get(url)
        resp.raise_for_status()
        log.info("Heruntergeladen: %d bytes", len(resp.content))
        return resp.content


def split_by_fonds(xlsx_bytes: bytes, fonds_col_idx: int, header_row_idx: int = 6) -> dict[str, bytes]:
    """Splittet die XLSX nach den Werten in der Fonds-Spalte. Gibt Dict
    {fonds_value: xlsx_bytes} zurueck. Header-Zeile (1-basiert: header_row_idx+1)
    bleibt in jedem Output erhalten.
    """
    from openpyxl import Workbook, load_workbook

    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=False, data_only=True)
    ws = wb.worksheets[0]
    all_rows = list(ws.iter_rows(values_only=True))
    header = all_rows[header_row_idx]
    data_rows = all_rows[header_row_idx + 2:]  # +2: header_row + en-Header

    groups: dict[str, list[tuple]] = {}
    for row in data_rows:
        if fonds_col_idx >= len(row):
            continue
        fonds = (row[fonds_col_idx] or "").strip() if row[fonds_col_idx] else ""
        if not fonds:
            continue
        groups.setdefault(fonds, []).append(row)

    output: dict[str, bytes] = {}
    for fonds, rows in groups.items():
        wb_out = Workbook()
        sheet = wb_out.active
        sheet.title = "Liste der Vorhaben"
        sheet.append(list(header))
        for r in rows:
            sheet.append(list(r))
        buf = io.BytesIO()
        wb_out.save(buf)
        output[fonds] = buf.getvalue()
        log.info("Split: %s -> %d Records", fonds, len(rows))
    return output


def upload(xlsx_bytes: bytes, filename: str, backend_url: str, worker_token: str, mode: str = "force") -> dict:
    log.info("Upload %s (mode=%s, %d bytes) -> %s", filename, mode, len(xlsx_bytes), backend_url)
    files = {"file": (filename, xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    data = {"mode": mode}
    headers = {"X-Worker-Token": worker_token}
    with httpx.Client(timeout=120.0) as client:
        resp = client.post(f"{backend_url}/api/beneficiaries/upload", files=files, data=data, headers=headers)
        resp.raise_for_status()
        result = resp.json()
        log.info(
            "Upload OK: source=%s rows=%d status=%s",
            result.get("source"), result.get("rows"), result.get("status"),
        )
        return result


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--url", required=True, help="Direkter Download-Link zur XLSX-Datei.")
    parser.add_argument(
        "--split-fonds-column",
        type=int,
        default=None,
        help=(
            "Wenn gesetzt: 0-basierter Spalten-Index der 'betroffener Fonds'-Spalte. "
            "Die Datei wird dann nach Fonds aufgeteilt und in mehreren Uploads "
            "an /upload geschickt. Brandenburg: 11."
        ),
    )
    parser.add_argument("--header-row", type=int, default=6, help="0-basierter Header-Row-Index (Brandenburg: 6).")
    parser.add_argument("--filename-prefix", default="harvested", help="Dateinamen-Praefix fuer den Upload.")
    parser.add_argument("--mode", default="force", choices=("smart", "full-refresh", "force"))
    parser.add_argument("--backend-url", default=os.getenv("BACKEND_URL", "http://localhost:8006"))
    parser.add_argument(
        "--worker-token",
        default=os.getenv("WORKER_API_TOKEN", "workshop-dev-auth-secret-change-me"),
    )
    parser.add_argument("--dry", action="store_true", help="Nur Download, kein Upload.")
    args = parser.parse_args()

    try:
        xlsx_bytes = download(args.url)
    except Exception as exc:  # noqa: BLE001
        log.error("Download fehlgeschlagen: %s", exc)
        return 2

    if args.split_fonds_column is not None:
        log.info("Splitten nach Fonds-Spalte %d", args.split_fonds_column)
        splits = split_by_fonds(xlsx_bytes, args.split_fonds_column, args.header_row)
        if args.dry:
            log.info("Dry-Run, Splits: %s", {k: f"{len(v)} bytes" for k, v in splits.items()})
            return 0
        for fonds, data in splits.items():
            fname = f"{args.filename_prefix}_{fonds.lower()}.xlsx"
            try:
                upload(data, fname, args.backend_url, args.worker_token, args.mode)
            except Exception as exc:  # noqa: BLE001
                log.error("Upload %s fehlgeschlagen: %s", fonds, exc)
                return 3
        return 0

    if args.dry:
        log.info("Dry-Run, %d bytes — kein Upload.", len(xlsx_bytes))
        return 0

    fname = f"{args.filename_prefix}.xlsx"
    try:
        upload(xlsx_bytes, fname, args.backend_url, args.worker_token, args.mode)
    except Exception as exc:  # noqa: BLE001
        log.error("Upload fehlgeschlagen: %s", exc)
        return 3
    return 0


if __name__ == "__main__":
    sys.exit(main())
