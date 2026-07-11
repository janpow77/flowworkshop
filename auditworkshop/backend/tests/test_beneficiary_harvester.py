"""
Unit-Tests fuer den Beneficiary-Harvester (Phase 6a).

Schwerpunkt:
  - compute_record_hash: deterministisch + stabil
  - parse_xlsx_or_csv: Mini-XLSX in-memory parsen, kanonische Felder finden
  - HarvestParams Default = smart

Wir testen rein die pure Logik (ohne DB), damit pytest schnell laeuft.
Der DB-seitige Smart/Force-Pfad ist analog zum State-Aid-Harvester
strukturiert und dort bereits durch Integrationstests abgedeckt.

Lauf: pytest backend/tests/test_beneficiary_harvester.py -q
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

# Backend-Verzeichnis in den Pfad legen, damit `services.*` importierbar ist.
_BACKEND = Path(__file__).resolve().parent.parent
if str(_BACKEND) not in sys.path:
    sys.path.insert(0, str(_BACKEND))


# ── compute_record_hash ──────────────────────────────────────────────────────


def test_compute_record_hash_deterministic_same_inputs():
    """Gleiche Inputs -> gleicher Hash, immer."""
    from services.beneficiary_harvester import compute_record_hash

    row = {
        "beneficiary_name": "Beispiel GmbH",
        "project_name": "Energieeffizienz 2030",
        "project_aktenzeichen": "AZ-12345",
        "bundesland": "Hessen",
        "periode": "2021-2027",
        "fonds": "EFRE",
        "funded_at_raw": "2024-03-15",
        "cost_total_raw": "150.000,00 €",
    }
    h1 = compute_record_hash(row, source_key="hessen_efre_2021_2027")
    h2 = compute_record_hash(row, source_key="hessen_efre_2021_2027")
    assert h1 == h2
    assert len(h1) == 32
    # Lower-Case Hex
    assert h1 == h1.lower()


def test_compute_record_hash_stable_across_whitespace():
    """Whitespace-Unterschiede aendern den Hash NICHT (Normalisierung greift)."""
    from services.beneficiary_harvester import compute_record_hash

    row_a = {"beneficiary_name": "Beispiel GmbH"}
    row_b = {"beneficiary_name": "  Beispiel   GmbH  "}
    assert (
        compute_record_hash(row_a, source_key="src1")
        == compute_record_hash(row_b, source_key="src1")
    )


def test_compute_record_hash_changes_with_source_key():
    """Verschiedene source_keys liefern verschiedene Hashes (sonst kollidieren
    Datensaetze unterschiedlicher Quellen)."""
    from services.beneficiary_harvester import compute_record_hash

    row = {"beneficiary_name": "Beispiel GmbH"}
    h_a = compute_record_hash(row, source_key="hessen_efre_2021_2027")
    h_b = compute_record_hash(row, source_key="bayern_efre_2021_2027")
    assert h_a != h_b


def test_compute_record_hash_changes_with_field_value():
    """Anderer Begünstigtenname → anderer Hash."""
    from services.beneficiary_harvester import compute_record_hash

    row_a = {"beneficiary_name": "Alpha GmbH"}
    row_b = {"beneficiary_name": "Beta GmbH"}
    assert (
        compute_record_hash(row_a, source_key="src1")
        != compute_record_hash(row_b, source_key="src1")
    )


def test_compute_record_hash_handles_none_values():
    """Felder mit None werden zu Leerstring normalisiert — kein Crash."""
    from services.beneficiary_harvester import compute_record_hash

    row = {
        "beneficiary_name": "X",
        "project_name": None,
        "cost_total_raw": None,
    }
    h = compute_record_hash(row, source_key="src1")
    assert isinstance(h, str)
    assert len(h) == 32


# ── parse_xlsx_or_csv ────────────────────────────────────────────────────────


def _build_minimal_xlsx() -> bytes:
    """Erzeugt eine Mini-XLSX in BytesIO mit den typischen kanonischen Spalten."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vorhaben"
    # Header (entspricht den in COLUMN_PATTERNS erkannten Patterns):
    headers = [
        "Name des Beguenstigten",   # → name
        "Bezeichnung des Vorhabens",  # → projekt
        "Foerderkennzeichen",       # → aktenzeichen
        "Zusammenfassung des Vorhabens",  # → beschreibung
        "Gesamtkosten des Vorhabens",  # → kosten
        "Projektstandort_PLZ",      # → plz
        "Projektstandort_Ort",      # → ort
        "Datum des Beginns des Vorhabens",  # → beginn
        "Datum des Endes des Vorhabens",    # → ende
    ]
    ws.append(headers)
    ws.append([
        "Acme GmbH", "Solaranlage Pilotprojekt", "AZ-001",
        "Photovoltaik Demonstrator", "150000",
        "60311", "Frankfurt am Main",
        "2024-03-01", "2026-12-31",
    ])
    ws.append([
        "Beta KG", "Energieeffiziente Sanierung", "AZ-002",
        "Modernisierung", "75000",
        "65183", "Wiesbaden",
        "2023-07-15", "2024-12-31",
    ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_xlsx_yields_rows_with_canonical_fields():
    """Mini-XLSX mit Standard-Headern -> kanonische Felder + raw_row."""
    from services.beneficiary_harvester import parse_xlsx_or_csv

    content = _build_minimal_xlsx()
    rows = list(parse_xlsx_or_csv(
        content, file_name="test.xlsx", sheet=0, header_row=0,
    ))
    # Pandas liest auch _row_number-Zeilen ohne Skip-Reason (die Header-Zeile
    # ist bereits konsumiert), also erwarten wir 2 Daten-Zeilen.
    valid = [r for r in rows if not r.get("_skip_reason")]
    assert len(valid) == 2

    first = valid[0]
    assert first["beneficiary_name"] == "Acme GmbH"
    assert first["project_name"] == "Solaranlage Pilotprojekt"
    assert first["project_aktenzeichen"] == "AZ-001"
    assert first["cost_total_raw"] == "150000"
    # Standort-Heuristik: COLUMN_PATTERNS hat `standort.*plz` vor `plz`,
    # daher kann Projektstandort_PLZ als location gelesen werden. Wir
    # akzeptieren beide Pfade — Hauptsache PLZ ist irgendwo erkannt.
    assert "60311" in (str(first.get("plz") or ""), str(first.get("location") or ""))

    # raw_row enthaelt die volle Original-Zeile
    assert "Name des Beguenstigten" in first["raw_row"]
    assert first["raw_row"]["Name des Beguenstigten"] == "Acme GmbH"


def test_parse_xlsx_skips_rows_without_name():
    """Zeilen ohne Begünstigtenname werden mit _skip_reason='no_name' geliefert."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Name des Beguenstigten", "Bezeichnung des Vorhabens"])
    ws.append([None, "Projekt ohne Name"])
    ws.append(["Valid GmbH", "Projekt mit Name"])
    buf = io.BytesIO()
    wb.save(buf)

    from services.beneficiary_harvester import parse_xlsx_or_csv
    rows = list(parse_xlsx_or_csv(
        buf.getvalue(), file_name="test.xlsx", sheet=0, header_row=0,
    ))
    skipped = [r for r in rows if r.get("_skip_reason") == "no_name"]
    valid = [r for r in rows if not r.get("_skip_reason")]
    assert len(skipped) == 1
    assert len(valid) == 1
    assert valid[0]["beneficiary_name"] == "Valid GmbH"


# ── HarvestParams Defaults ────────────────────────────────────────────────────


def test_harvest_params_default_mode_is_snapshot():
    """Ohne explizite mode-Angabe: vollständiger, validierter Quellensnapshot."""
    from services.beneficiary_harvester import BeneficiaryHarvestParams

    p = BeneficiaryHarvestParams(source_key="src1")
    assert p.mode == "snapshot"


def test_harvest_params_accepts_known_modes():
    """Smart, full-refresh, force sind alle akzeptiert."""
    from services.beneficiary_harvester import BeneficiaryHarvestParams

    for mode in ("snapshot", "smart", "full-refresh", "force"):
        p = BeneficiaryHarvestParams(source_key="src1", mode=mode)
        assert p.mode == mode


def test_snapshot_validation_rejects_invalid_financial_and_date_relations():
    from services.beneficiary_harvester import BeneficiaryHarvestParams, validate_beneficiary_rows
    params = BeneficiaryHarvestParams(
        source_key="src", fonds="EFRE", periode="2021-2027", country_code="DE",
    )
    errors = validate_beneficiary_rows([{
        "_row_number": 4, "beneficiary_name": "Beispiel GmbH",
        "cost_total_raw": "100", "cost_eu_funding_raw": "120",
        "project_start_raw": "2025-12-31", "project_end_raw": "2025-01-01",
        "latitude": 100, "longitude": 200,
    }], params)
    assert len(errors) == 4


# ── _detect_canonical_columns: explicit_mapping hat Vorrang ──────────────────


def test_detect_canonical_columns_explicit_overrides_pattern():
    """Wenn der Aufrufer ein Mapping mitgibt, gilt das vor den Patterns."""
    from services.beneficiary_harvester import _detect_canonical_columns

    headers = ["Foo", "Bar", "Name des Beguenstigten"]
    mapping = _detect_canonical_columns(
        headers, explicit_mapping={"name": "Foo"},
    )
    assert mapping["name"] == "Foo"


def test_detect_canonical_columns_pattern_fallback():
    """Ohne explicit_mapping fallen wir auf die COLUMN_PATTERNS-Regex zurueck."""
    from services.beneficiary_harvester import _detect_canonical_columns

    headers = [
        "Name des Beguenstigten", "Bezeichnung des Vorhabens",
        "Gesamtkosten des Vorhabens",
    ]
    mapping = _detect_canonical_columns(headers, explicit_mapping=None)
    assert "name" in mapping
    assert mapping["name"] == "Name des Beguenstigten"
    assert "projekt" in mapping
    assert "kosten" in mapping


# ── Smart-Mode-Integration (DB) ──────────────────────────────────────────────


def _has_db_connection() -> bool:
    """True, wenn die zentrale Tabelle ueber die echte Engine erreichbar ist.

    Reine Unit-Tests laufen ohne Docker — dort skippen wir die Integration.
    Innerhalb des Containers (oder mit DATABASE_URL gesetzt) laufen sie
    durch.
    """
    try:
        from sqlalchemy import inspect
        from database import engine
        return inspect(engine).has_table("workshop_beneficiary_records")
    except Exception:
        return False


def _seed_xlsx() -> bytes:
    """XLSX mit einer minimalen Beneficiary-Zeile fuer Smart-Mode-Tests."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "Name des Beguenstigten",
        "Bezeichnung des Vorhabens",
        "Gesamtkosten des Vorhabens",
    ])
    ws.append(["Smart Test GmbH", "Workshop Demo", "100000"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_smart_mode_second_run_skips_existing():
    """Phase 6a: Zweiter Smart-Lauf mit identischen Daten → 0 inserted."""
    import pytest

    if not _has_db_connection():
        pytest.skip("workshop_beneficiary_records-Tabelle nicht erreichbar.")

    from database import SessionLocal
    from services.beneficiary_harvester import (
        BeneficiaryHarvestParams, run_beneficiary_harvest,
    )
    from models.beneficiary_records import BeneficiaryRecord

    db = SessionLocal()
    test_source = "test_smart_mode_phase6a"
    try:
        # Aufraeumen — Test ist idempotent.
        db.query(BeneficiaryRecord).filter(
            BeneficiaryRecord.source_key == test_source
        ).delete(synchronize_session=False)
        db.commit()

        content = _seed_xlsx()
        params = BeneficiaryHarvestParams(
            source_key=test_source,
            file_content=content,
            file_name="smart_test.xlsx",
            mode="smart",
            triggered_by="pytest",
        )

        result_a = run_beneficiary_harvest(db, params)
        assert result_a["status"] in ("ok", "partial")
        assert result_a["records_inserted"] == 1
        assert result_a["records_skipped"] == 0

        # Zweiter Lauf — selbe Daten, smart-Mode: 0 neu, 1 skip.
        result_b = run_beneficiary_harvest(db, params)
        assert result_b["records_inserted"] == 0
        assert result_b["records_skipped"] == 1
    finally:
        # Aufraeumen (Test-Daten loeschen).
        db.query(BeneficiaryRecord).filter(
            BeneficiaryRecord.source_key == test_source
        ).delete(synchronize_session=False)
        db.commit()
        db.close()


def test_force_mode_predeletes_records():
    """Phase 6a: force-Mode loescht den Bestand der Quelle vor dem Insert."""
    import pytest

    if not _has_db_connection():
        pytest.skip("workshop_beneficiary_records-Tabelle nicht erreichbar.")

    from database import SessionLocal
    from services.beneficiary_harvester import (
        BeneficiaryHarvestParams, run_beneficiary_harvest,
    )
    from models.beneficiary_records import BeneficiaryRecord

    db = SessionLocal()
    test_source = "test_force_mode_phase6a"
    try:
        # Erst-Befuellung.
        params = BeneficiaryHarvestParams(
            source_key=test_source,
            file_content=_seed_xlsx(),
            file_name="force_test.xlsx",
            mode="smart",
            triggered_by="pytest",
        )
        run_beneficiary_harvest(db, params)

        before = (
            db.query(BeneficiaryRecord)
            .filter(BeneficiaryRecord.source_key == test_source)
            .count()
        )
        assert before == 1

        # Force-Lauf — Pre-Delete + reiner Insert.
        params.mode = "force"
        result = run_beneficiary_harvest(db, params)
        assert result["records_inserted"] == 1

        after = (
            db.query(BeneficiaryRecord)
            .filter(BeneficiaryRecord.source_key == test_source)
            .count()
        )
        # Genau 1 — der vorhergehende wurde geloescht und neu eingefuegt.
        assert after == 1
    finally:
        db.query(BeneficiaryRecord).filter(
            BeneficiaryRecord.source_key == test_source
        ).delete(synchronize_session=False)
        db.commit()
        db.close()


def test_full_refresh_updates_existing():
    """Phase 6a: full-refresh aktualisiert Felder bei Hash-Konflikt."""
    import pytest
    import openpyxl

    if not _has_db_connection():
        pytest.skip("workshop_beneficiary_records-Tabelle nicht erreichbar.")

    from database import SessionLocal
    from services.beneficiary_harvester import (
        BeneficiaryHarvestParams, run_beneficiary_harvest,
    )
    from models.beneficiary_records import BeneficiaryRecord

    db = SessionLocal()
    test_source = "test_full_refresh_phase6a"
    try:
        # Erstlauf — Project-Beschreibung mit "zweck" matcht beschreibung-Pattern.
        # Hash-relevante Felder (name + projekt + bundesland + ...) bleiben gleich,
        # nur die Beschreibung aendert sich → full-refresh muss das Update
        # durchziehen.
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([
            "Name des Beguenstigten",
            "Bezeichnung des Vorhabens",
            "Zweck des Vorhabens",
            "Gesamtkosten des Vorhabens",
        ])
        ws.append(["Refresh GmbH", "Demo", "v1", "100000"])
        buf_a = io.BytesIO()
        wb.save(buf_a)

        params_a = BeneficiaryHarvestParams(
            source_key=test_source,
            file_content=buf_a.getvalue(),
            file_name="refresh_a.xlsx",
            mode="smart",
            triggered_by="pytest",
        )
        run_beneficiary_harvest(db, params_a)

        # Zweite Datei — gleiche Hash-Felder, andere Description.
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2.append([
            "Name des Beguenstigten",
            "Bezeichnung des Vorhabens",
            "Zweck des Vorhabens",
            "Gesamtkosten des Vorhabens",
        ])
        ws2.append(["Refresh GmbH", "Demo", "v2-updated", "100000"])
        buf_b = io.BytesIO()
        wb2.save(buf_b)

        params_b = BeneficiaryHarvestParams(
            source_key=test_source,
            file_content=buf_b.getvalue(),
            file_name="refresh_b.xlsx",
            mode="full-refresh",
            triggered_by="pytest",
        )
        run_beneficiary_harvest(db, params_b)

        rec = (
            db.query(BeneficiaryRecord)
            .filter(BeneficiaryRecord.source_key == test_source)
            .first()
        )
        assert rec is not None
        assert rec.project_description == "v2-updated"
    finally:
        db.query(BeneficiaryRecord).filter(
            BeneficiaryRecord.source_key == test_source
        ).delete(synchronize_session=False)
        db.commit()
        db.close()
